from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from arxml_codegen.contract.schema import DeliveryContract, RunnableContract
from arxml_codegen.excel.template import SHEETS


def write_contract_excel(contract: DeliveryContract, output: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    rows_by_sheet = build_workbook_rows(contract)

    for sheet_name, headers in SHEETS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for row in rows_by_sheet.get(sheet_name, []):
            sheet.append([row.get(header, "") for header in headers])
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column) + 2
            sheet.column_dimensions[column[0].column_letter].width = min(max(width, 14), 64)

    _add_validations(workbook)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def build_workbook_rows(contract: DeliveryContract) -> dict[str, list[dict[str, str]]]:
    profile = _profile(contract)
    if profile == "signal_atomic_davinci":
        return _build_signal_atomic_davinci_rows(contract)
    if profile == "mixed_signal_soa":
        return _build_mixed_signal_soa_rows(contract)

    return _build_generic_rows(contract)


def _build_generic_rows(contract: DeliveryContract) -> dict[str, list[dict[str, str]]]:

    root = _root(contract)
    paths = _davinci_paths(contract)
    mapping_path = paths["mapping_set_path"]
    has_explicit_connectors = any(
        connector.provider_endpoint or connector.requester_endpoint
        for connector in contract.connectors
    )
    rows: dict[str, list[dict[str, str]]] = {
        "ProjectConfig": [
            {"Key": "AutosarVersion", "Value": contract.project.target_autosar_version},
            {"Key": "RootPackage", "Value": root},
            {"Key": "DefaultMappingSetPath", "Value": mapping_path},
            {"Key": "InterfacePackage", "Value": paths["interface_package"]},
            {"Key": "DataTypePackage", "Value": paths["data_type_package"]},
            {"Key": "CompuMethodPackage", "Value": paths["compu_method_package"]},
            {"Key": "DataConstrPackage", "Value": paths["data_constr_package"]},
            {"Key": "UnitPackage", "Value": paths["unit_package"]},
        ],
        "Components": [],
        "ComponentPrototypes": [],
        "PrimitiveDataTypes": [],
        "RecordTypes": [],
        "RecordElements": [],
        "PortRecordInitValues": [],
        "DataTypeMappings": [],
        "CompuMethods": [],
        "CompuScales": [],
        "DataConstrs": [],
        "SRInterfaces": [],
        "SRDataElements": [],
        "Ports": [],
        "Runnables": [],
        "RunnableEvents": [],
        "RunnableAccesses": [],
        "CSInterfaces": [],
        "CSOperations": [],
        "CSArguments": [],
        "CompositionConnectors": [],
        "Units": [],
    }

    composition_name = _composition_name(contract)
    if composition_name and not any(swc.name == composition_name for swc in contract.swcs):
        rows["Components"].append(
            {
                "ComponentName": composition_name,
                "ComponentKind": "Composition",
                "PackagePath": f"{root}/Components",
                "Description": "Inferred top-level composition for generated connectors",
            }
        )

    for swc in contract.swcs:
        if not swc.name:
            continue
        rows["Components"].append(
            {
                "ComponentName": swc.name,
                "ComponentKind": swc.kind or "Application",
                "PackagePath": f"{root}/Components",
                "InternalBehaviorName": f"IB_{swc.name}",
                "ImplementationName": f"{swc.name}_Impl",
                "Description": swc.description,
            }
        )
        if (swc.kind or "Application") != "Composition" and composition_name:
            rows["ComponentPrototypes"].append(
                {
                    "CompositionName": composition_name,
                    "PrototypeName": _prototype_for_component(contract, swc.name),
                    "ComponentTypeName": swc.name,
                    "ComponentTypeRef": f"{root}/Components/{swc.name}",
                    "Description": swc.description,
                }
            )

    for data_type in contract.data_types:
        type_name = _app_type(data_type.type_name)
        app_ref = _data_type_ref(paths, type_name)
        impl_ref = f"/AUTOSAR_Platform/ImplementationDataTypes/{data_type.base_type or 'uint8'}"
        compu_ref = ""
        constr_ref = ""
        if data_type.enum_values:
            compu_ref = f'{paths["compu_method_package"]}/CM_{type_name}_TextTable'
            rows["CompuMethods"].append(
                {
                    "CompuMethodName": f"CM_{type_name}_TextTable",
                    "CompuMethodPath": compu_ref,
                    "Category": "TEXTTABLE",
                    "Description": data_type.description,
                }
            )
            for lower, upper, enum_text in _split_enum_scales(data_type.enum_values):
                rows["CompuScales"].append(
                    {
                        "CompuMethodName": f"CM_{type_name}_TextTable",
                        "LowerLimit": lower,
                        "UpperLimit": upper,
                        "TextValue": enum_text,
                    }
                )
        if data_type.physical_range:
            lower, upper = _split_range(data_type.physical_range)
            constr_ref = f'{paths["data_constr_package"]}/DC_{type_name}'
            rows["DataConstrs"].append(
                {
                    "DataConstrName": f"DC_{type_name}",
                    "DataConstrPath": constr_ref,
                    "LowerLimit": lower,
                    "UpperLimit": upper,
                    "Description": data_type.description,
                }
            )
        rows["PrimitiveDataTypes"].append(
            {
                "ApplicationTypeName": type_name,
                "ApplicationTypePath": app_ref,
                "ImplementationTypeName": data_type.base_type or "uint8",
                "ImplementationTypePath": impl_ref,
                "BaseType": data_type.base_type or "uint8",
                "CompuMethodRef": compu_ref,
                "DataConstrRef": constr_ref,
                "CalibrationAccess": "READ-ONLY",
                "UnitRef": _unit_ref(paths, data_type.unit),
                "Description": data_type.description,
            }
        )
        rows["DataTypeMappings"].append(
            {
                "MappingSetPath": mapping_path,
                "ApplicationTypeRef": app_ref,
                "ImplementationTypeRef": impl_ref,
            }
        )

    record_names = sorted({element.record_type for element in contract.record_elements if element.record_type})
    for record_name in record_names:
        app_type = _app_type(record_name)
        impl_type = f"Impl_{_short(record_name)}"
        rows["RecordTypes"].append(
            {
                "ApplicationTypeName": app_type,
                "ApplicationTypePath": _data_type_ref(paths, app_type),
                "ImplementationTypeName": impl_type,
                "ImplementationTypePath": _data_type_ref(paths, impl_type),
                "CalibrationAccess": "READ-ONLY",
                "Description": "Derived from delivery contract record definition",
            }
        )
        rows["DataTypeMappings"].append(
            {
                "MappingSetPath": mapping_path,
                "ApplicationTypeRef": _data_type_ref(paths, app_type),
                "ImplementationTypeRef": _data_type_ref(paths, impl_type),
            }
        )

    for order, element in enumerate(contract.record_elements, start=1):
        if not element.record_type or not element.element_name:
            continue
        is_record = (element.field_category or "").strip().lower() == "record"
        implementation_type = _short(element.implementation_field_type)
        rows["RecordElements"].append(
            {
                "RecordTypeName": _app_type(element.record_type),
                "ElementName": element.element_name,
                "ImplementationElementName": element.implementation_element_name or element.element_name,
                "ElementCategory": "Record" if is_record else element.field_category,
                "ApplicationElementTypeRef": _data_type_ref(paths, _app_type(element.data_type)),
                "ImplementationElementTypeRef": (
                    _data_type_ref(paths, implementation_type)
                    if is_record
                    else f"/AUTOSAR_Platform/ImplementationDataTypes/{_base_type_for_contract(contract, element.implementation_field_type or element.data_type)}"
                ),
                "Order": str(order),
                "Description": element.description,
            }
        )

    units = sorted({dt.unit for dt in contract.data_types if dt.unit})
    rows["Units"] = [
        {
            "UnitName": f"Unit_{_short(unit)}",
            "UnitPath": _unit_ref(paths, unit),
            "DisplayName": unit,
            "FactorSIToUnit": "1",
            "OffsetSIToUnit": "0",
        }
        for unit in units
    ]

    for signal in contract.signals:
        if not signal.signal_name:
            continue
        interface_name = f"If_{signal.signal_name}_SR"
        interface_path = _interface_ref(paths, interface_name)
        data_element = f"DE_{signal.signal_name}"
        app_ref = _data_type_ref(paths, _app_type(signal.data_type))
        rows["SRInterfaces"].append(
            {
                "InterfaceName": interface_name,
                "InterfacePath": interface_path,
                "IsService": "false",
                "Description": signal.description,
            }
        )
        rows["SRDataElements"].append(
            {
                "InterfaceName": interface_name,
                "DataElementName": data_element,
                "ApplicationTypeRef": app_ref,
                "Description": signal.description,
            }
        )
        if signal.provider_swc:
            rows["Ports"].append(
                _sr_port_row(signal.provider_swc, f"Pp_{signal.signal_name}", "P", interface_path, data_element, signal)
            )
        if signal.consumer_swc:
            rows["Ports"].append(
                _sr_port_row(signal.consumer_swc, f"Rp_{signal.signal_name}", "R", interface_path, data_element, signal)
            )
        if not has_explicit_connectors and composition_name and signal.provider_swc and signal.consumer_swc:
            rows["CompositionConnectors"].append(
                _connector_row(
                    contract,
                    composition_name,
                    signal.provider_swc,
                    f"Pp_{signal.signal_name}",
                    signal.consumer_swc,
                    f"Rp_{signal.signal_name}",
                    f"SR: {signal.signal_name}",
                )
            )

    for service in contract.services:
        interface_name = service.interface_name or f"If_{service.service_name}_CS"
        interface_path = _interface_ref(paths, interface_name)
        operation_name = service.operation_name or service.service_name
        rows["CSInterfaces"].append(
            {"InterfaceName": interface_name, "InterfacePath": interface_path, "IsService": "false"}
        )
        rows["CSOperations"].append(
            {"InterfaceName": interface_name, "OperationName": operation_name, "Description": service.description}
        )
        if service.owner_swc and service.port_name:
            direction = _port_direction_from_role(service.port_role)
            if direction:
                rows["Ports"].append(
                    _cs_port_row(service.owner_swc, service.port_name, direction, interface_path, operation_name)
                )
        elif service.provider_swc:
            rows["Ports"].append(
                _cs_port_row(service.provider_swc, f"Pp_{service.service_name}", "P", interface_path, operation_name)
            )
        if not service.owner_swc and service.client_swc:
            rows["Ports"].append(
                _cs_port_row(service.client_swc, f"Rp_{service.service_name}", "R", interface_path, operation_name)
            )
        if not has_explicit_connectors and composition_name and service.provider_swc and service.client_swc:
            rows["CompositionConnectors"].append(
                _connector_row(
                    contract,
                    composition_name,
                    service.provider_swc,
                    f"Pp_{service.service_name}",
                    service.client_swc,
                    f"Rp_{service.service_name}",
                    f"CS: {operation_name}",
                )
            )

    for arg in contract.operation_args:
        if not arg.operation_name or arg.direction.upper() not in {"IN", "OUT", "INOUT"}:
            continue
        rows["CSArguments"].append(
            {
                "InterfaceName": _find_interface_for_operation(
                    contract,
                    arg.operation_name,
                    arg.interface_name,
                ),
                "OperationName": arg.operation_name,
                "ArgumentName": arg.argument_name,
                "Direction": arg.direction,
                "ApplicationTypeRef": _data_type_ref(paths, _app_type(arg.data_type)),
                "Description": arg.description,
            }
        )

    for connector in contract.connectors:
        row = _connector_row_from_contract(contract, composition_name, connector)
        if row:
            rows["CompositionConnectors"].append(row)

    operation_bindings = _operation_event_bindings(contract)
    for runnable in contract.runnables:
        if not runnable.swc or not runnable.runnable_name:
            continue
        rows["Runnables"].append(
            {
                "ComponentName": runnable.swc,
                "RunnableName": runnable.runnable_name,
                "Symbol": runnable.runnable_name,
                "Description": runnable.description,
            }
        )
        binding = operation_bindings.get((runnable.swc, runnable.runnable_name), {}) if runnable.trigger_type == "OperationInvoked" else {}
        if runnable.trigger_type == "OperationInvoked" and not binding:
            binding = _infer_operation_event_binding(contract, runnable)
        related_port = (
            _related_port(binding.get("port", ""))
            or _related_port(runnable.related_port_or_signal)
            or _infer_operation_port(contract, runnable)
        )
        related_operation = binding.get("operation", "") or runnable.related_operation
        if runnable.trigger_type and (runnable.trigger_type != "OperationInvoked" or related_port):
            rows["RunnableEvents"].append(
                {
                    "ComponentName": runnable.swc,
                    "RunnableName": runnable.runnable_name,
                    "TriggerType": runnable.trigger_type,
                    "PeriodMs": runnable.period_ms,
                    "PortName": related_port,
                    "OperationName": related_operation,
                    "DataElementName": _related_data_element(runnable.related_port_or_signal),
                }
            )
        for signal_name in _split_list(runnable.read_signals):
            port_name = _access_port_name(signal_name, "R")
            rows["RunnableAccesses"].append(
                _runnable_access_row(
                    runnable.swc,
                    runnable.runnable_name,
                    "DataRead",
                    port_name,
                    "",
                    _data_element_from_port_or_signal(signal_name),
                )
            )
        for signal_name in _split_list(runnable.write_signals):
            port_name = _access_port_name(signal_name, "P")
            rows["RunnableAccesses"].append(
                _runnable_access_row(
                    runnable.swc,
                    runnable.runnable_name,
                    "DataWrite",
                    port_name,
                    "",
                    _data_element_from_port_or_signal(signal_name),
                )
            )
        if related_operation and _is_client_service_port(
            contract,
            runnable.swc,
            related_port,
            related_operation,
        ):
            rows["RunnableAccesses"].append(
                _runnable_access_row(
                    runnable.swc,
                    runnable.runnable_name,
                    "ServerCallPoint",
                    related_port,
                    related_operation,
                    "",
                )
            )

    _dedupe_rows(rows)
    return rows


def _build_mixed_signal_soa_rows(contract: DeliveryContract) -> dict[str, list[dict[str, str]]]:
    """Build the multi-SWC CP model while preserving DaVinci signal naming.

    Signal tables define the reusable S/R interface and data-type pool. Actual
    component S/R ports are created only from Runnable Access rows. C/S ports,
    operations, composition prototypes and connectors continue to use the
    generic SOA builder.
    """
    rows = _build_generic_rows(contract)
    signal_rows = _build_signal_atomic_davinci_rows(contract)
    mapping_path = "/ComponentTypes/MappingSets/APP_data_mapping"

    # Reuse the proven signal data-type rules: Record and Primitive types are
    # mutually exclusive, Boolean uses the platform CM/DC, and LINEAR keeps the
    # document's physical limits, resolution and offset.
    data_type_sheets = (
        "PrimitiveDataTypes",
        "RecordTypes",
        "RecordElements",
        "DataTypeMappings",
        "CompuMethods",
        "CompuScales",
        "DataConstrs",
        "Units",
    )
    for sheet in data_type_sheets:
        rows[sheet] = [dict(row) for row in signal_rows[sheet]]
    for row in rows["DataTypeMappings"]:
        row["MappingSetPath"] = mapping_path

    # DaVinci's reusable package layout places SWC types directly below
    # /ComponentTypes; interfaces and data types remain in their own packages.
    for row in rows["ProjectConfig"]:
        if row.get("Key") == "RootPackage":
            row["Value"] = "/ComponentTypes"
        elif row.get("Key") == "DefaultMappingSetPath":
            row["Value"] = mapping_path
    rows["ProjectConfig"].append({"Key": "GenerationProfile", "Value": "mixed_signal_soa"})
    for row in rows["Components"]:
        row["PackagePath"] = "/ComponentTypes"
    for row in rows["ComponentPrototypes"]:
        component_name = row.get("ComponentTypeName", "")
        row["ComponentTypeRef"] = f"/ComponentTypes/{component_name}"

    # The signal tables configure only Application Port Interfaces. Interface
    # and DataElement SHORT-NAMEs both remain exactly equal to SignalName.
    rows["SRInterfaces"] = []
    rows["SRDataElements"] = []
    signals_by_name = {}
    for signal in contract.signals:
        signal_name = _short(signal.signal_name)
        if not signal_name:
            continue
        signals_by_name[signal_name] = signal
        rows["SRInterfaces"].append(
            {
                "InterfaceName": signal_name,
                "InterfacePath": f"/PortInterfaces/{signal_name}",
                "IsService": "false",
                "Description": signal.description,
            }
        )
        rows["SRDataElements"].append(
            {
                "InterfaceName": signal_name,
                "DataElementName": signal_name,
                "ApplicationTypeRef": _signal_app_type_ref(
                    signal, _base_type_for_signal(contract, signal)
                ),
                "Description": signal.description,
            }
        )

    # Preserve explicit C/S ports. S/R component ports and accesses are rebuilt
    # solely from Runnable Access, using its signal/port name verbatim.
    rows["Ports"] = [row for row in rows["Ports"] if row.get("InterfaceKind") != "SR"]
    rows["RunnableAccesses"] = [
        row for row in rows["RunnableAccesses"] if row.get("AccessType") == "ServerCallPoint"
    ]
    rows["PortRecordInitValues"] = []
    sr_ports_added: set[tuple[str, str]] = set()
    for runnable in contract.runnables:
        if not runnable.swc or not runnable.runnable_name:
            continue
        for access_type, direction, values in (
            ("DataRead", "R", runnable.read_signals),
            ("DataWrite", "P", runnable.write_signals),
        ):
            for raw_name in _split_list(values):
                port_name = _short(raw_name)
                signal = signals_by_name.get(port_name)
                if signal and (runnable.swc, port_name) not in sr_ports_added:
                    port_row = _sr_port_row(
                        runnable.swc,
                        port_name,
                        direction,
                        f"/PortInterfaces/{port_name}",
                        port_name,
                        signal,
                    )
                    base_type = _base_type_for_signal(contract, signal)
                    port_row["InitValue"] = _signal_init_value(signal, base_type)
                    port_row["InitValueType"] = _signal_init_value_type(signal, base_type)
                    rows["Ports"].append(port_row)
                    if port_row["InitValueType"] == "Record":
                        _add_record_init_rows(rows, runnable.swc, port_name, signal, contract)
                    sr_ports_added.add((runnable.swc, port_name))

                # Keep a missing signal reference visible to model validation;
                # the DOCX gap report already explains that it must be defined.
                rows["RunnableAccesses"].append(
                    _runnable_access_row(
                        runnable.swc,
                        runnable.runnable_name,
                        access_type,
                        port_name,
                        "",
                        port_name,
                    )
                )

    _dedupe_rows(rows)
    return rows


def _build_signal_atomic_davinci_rows(contract: DeliveryContract) -> dict[str, list[dict[str, str]]]:
    swc_name = _atomic_swc_name(contract)
    mapping_path = "/ComponentTypes/MappingSets/DataMapping"
    rows: dict[str, list[dict[str, str]]] = {
        "ProjectConfig": [
            {"Key": "AutosarVersion", "Value": contract.project.target_autosar_version},
            {"Key": "RootPackage", "Value": "/ComponentTypes"},
            {"Key": "GenerationProfile", "Value": "signal_atomic_davinci"},
            {"Key": "DefaultMappingSetPath", "Value": mapping_path},
            {"Key": "InterfacePackage", "Value": "/PortInterfaces"},
            {"Key": "DataTypePackage", "Value": "/DataTypes"},
            {"Key": "UnitPackage", "Value": "/DataTypes/Units"},
        ],
        "Components": [
            {
                "ComponentName": swc_name,
                "ComponentKind": "Application",
                "PackagePath": "/ComponentTypes",
                "InternalBehaviorName": f"{swc_name}_InternalBehavior",
                "ImplementationName": f"{swc_name}_Impl",
                "Description": "Single atomic SWC generated from signal delivery document",
            }
        ],
        "ComponentPrototypes": [],
        "PrimitiveDataTypes": [],
        "RecordTypes": [],
        "RecordElements": [],
        "PortRecordInitValues": [],
        "DataTypeMappings": [],
        "CompuMethods": [],
        "CompuScales": [],
        "DataConstrs": [],
        "SRInterfaces": [],
        "SRDataElements": [],
        "Ports": [],
        "Runnables": [],
        "RunnableEvents": [],
        "RunnableAccesses": [],
        "CSInterfaces": [],
        "CSOperations": [],
        "CSArguments": [],
        "CompositionConnectors": [],
        "Units": [
            {
                "UnitName": "No_Unit",
                "UnitPath": "/DataTypes/Units/No_Unit",
                "DisplayName": "-",
                "FactorSIToUnit": "1",
                "OffsetSIToUnit": "0",
            }
        ],
    }

    used_base_types = sorted(
        {
            _base_type_for_signal(contract, signal)
            for signal in contract.signals
            if signal.signal_name
        }
        | {
            _normalize_base_type(data_type.base_type)
            for data_type in contract.data_types
            if (data_type.type_kind or "").strip().lower() != "record"
        }
        | {
            _normalize_base_type(element.implementation_field_type or element.data_type)
            for element in contract.record_elements
            if (element.field_category or "").strip().lower() != "record"
        }
        | {
            _normalize_base_type(argument.internal_data_type or argument.data_type)
            for argument in contract.operation_args
            if (argument.value_type or "").strip().lower() != "record"
        }
    )
    for base_type in used_base_types:
        app_type = f"App_{base_type}"
        app_ref = f"/DataTypes/{app_type}"
        impl_ref = f"/AUTOSAR_Platform/ImplementationDataTypes/{base_type}"
        constr_ref = f"/AUTOSAR_Platform/DataConstrs/{base_type}_DataConstr"
        compu_method = _base_type_compu_method_name(base_type)
        compu_ref = _base_type_compu_method_ref(base_type)
        rows["PrimitiveDataTypes"].append(
            {
                "ApplicationTypeName": app_type,
                "ApplicationTypePath": app_ref,
                "ImplementationTypeName": base_type,
                "ImplementationTypePath": impl_ref,
                "BaseType": base_type,
                "CompuMethodRef": compu_ref,
                "DataConstrRef": constr_ref,
                "CalibrationAccess": "READ-ONLY",
                "UnitRef": _shared_base_type_unit_ref(contract, base_type),
                "Description": f"Shared application type for {base_type} signals",
            }
        )
        rows["CompuMethods"].append(
            {
                "CompuMethodName": compu_method,
                "CompuMethodPath": compu_ref,
                "Category": "TEXTTABLE" if base_type == "boolean" else "IDENTICAL",
                "Description": f"Shared {base_type} conversion method",
            }
        )
        rows["DataConstrs"].append(
            {
                "DataConstrName": f"{base_type}_DataConstr",
                "DataConstrPath": constr_ref,
                "LowerLimit": "0",
                "UpperLimit": _base_type_upper(base_type),
                "Description": f"Platform {base_type} internal constraint",
            }
        )
        rows["DataTypeMappings"].append(
            {
                "MappingSetPath": mapping_path,
                "ApplicationTypeRef": app_ref,
                "ImplementationTypeRef": impl_ref,
            }
        )

    if "boolean" in used_base_types:
        rows["CompuScales"].extend(
            [
                {"CompuMethodName": "boolean_CompuMethod", "LowerLimit": "0", "UpperLimit": "0", "TextValue": "false"},
                {"CompuMethodName": "boolean_CompuMethod", "LowerLimit": "1", "UpperLimit": "1", "TextValue": "true"},
            ]
        )

    for unit in sorted(
        {
            dt.unit
            for dt in contract.data_types
            if dt.unit
            and _short(dt.unit)
            and _short(dt.unit).lower() != "no_unit"
        }
    ):
        rows["Units"].append(
            {
                "UnitName": f"Unit_{_short(unit)}",
                "UnitPath": f"/DataTypes/Units/Unit_{_short(unit)}",
                "DisplayName": unit,
                "FactorSIToUnit": "1",
                "OffsetSIToUnit": "0",
            }
        )

    for data_type in contract.data_types:
        _add_signal_atomic_data_type_rows(rows, data_type, mapping_path)

    for record_type in _signal_atomic_record_types(contract):
        app_type = _short(record_type["app"])
        impl_type = _short(record_type["impl"])
        rows["RecordTypes"].append(
            {
                "ApplicationTypeName": app_type,
                "ApplicationTypePath": f"/DataTypes/{app_type}",
                "ImplementationTypeName": impl_type,
                "ImplementationTypePath": f"/DataTypes/{impl_type}",
                "CalibrationAccess": "READ-ONLY",
                "Description": "Derived from signal delivery record definition",
            }
        )
        rows["DataTypeMappings"].append(
            {
                "MappingSetPath": mapping_path,
                "ApplicationTypeRef": f"/DataTypes/{app_type}",
                "ImplementationTypeRef": f"/DataTypes/{impl_type}",
            }
        )

    for element in sorted(contract.record_elements, key=lambda item: int(item.field_order or 0)):
        if not element.record_type or not element.element_name:
            continue
        app_type = _short(element.record_type)
        field_app = _short(element.data_type)
        is_record = (element.field_category or "").strip().lower() == "record"
        impl_field = _short(element.implementation_field_type) if is_record else _normalize_base_type(element.implementation_field_type or element.data_type)
        rows["RecordElements"].append(
            {
                "RecordTypeName": app_type,
                "ElementName": element.element_name,
                "ImplementationElementName": element.implementation_element_name or element.element_name,
                "ElementCategory": "Record" if is_record else element.field_category,
                "ApplicationElementTypeRef": f"/DataTypes/{field_app}",
                "ImplementationElementTypeRef": (
                    f"/DataTypes/{impl_field}"
                    if is_record
                    else f"/AUTOSAR_Platform/ImplementationDataTypes/{impl_field}"
                ),
                "Order": element.field_order,
                "Description": element.description,
            }
        )

    for signal in contract.signals:
        if not signal.signal_name:
            continue
        signal_name = _short(signal.signal_name)
        base_type = _base_type_for_signal(contract, signal)
        app_ref = _signal_app_type_ref(signal, base_type)
        interface_path = f"/PortInterfaces/{signal_name}"
        rows["SRInterfaces"].append(
            {
                "InterfaceName": signal_name,
                "InterfacePath": interface_path,
                "IsService": "false",
                "Description": signal.description,
            }
        )
        rows["SRDataElements"].append(
            {
                "InterfaceName": signal_name,
                "DataElementName": signal_name,
                "ApplicationTypeRef": app_ref,
                "Description": signal.description,
            }
        )

    operation_event_bindings = _operation_event_bindings(contract)
    signals_by_name = {_short(signal.signal_name): signal for signal in contract.signals}
    ports_added: set[tuple[str, str]] = set()

    for runnable in contract.runnables:
        if not runnable.runnable_name:
            continue
        rows["Runnables"].append(
            {
                "ComponentName": swc_name,
                "RunnableName": runnable.runnable_name,
                "Symbol": runnable.runnable_name,
                "Description": runnable.description,
            }
        )
        if runnable.trigger_type:
            binding = operation_event_bindings.get((runnable.swc, runnable.runnable_name), {})
            port_name = runnable.related_port_or_signal or binding.get("port", "")
            operation_name = runnable.related_operation or binding.get("operation", "")
            rows["RunnableEvents"].append(
                {
                    "ComponentName": swc_name,
                    "RunnableName": runnable.runnable_name,
                    "TriggerType": runnable.trigger_type,
                    "PeriodMs": runnable.period_ms,
                    "PortName": port_name,
                    "OperationName": operation_name,
                }
            )
        elif runnable.related_operation and runnable.related_port_or_signal:
            rows["RunnableEvents"].append(
                {
                    "ComponentName": swc_name,
                    "RunnableName": runnable.runnable_name,
                    "TriggerType": "OperationInvoked",
                    "PortName": runnable.related_port_or_signal,
                    "OperationName": runnable.related_operation,
                }
            )
        for signal_name in _split_list(runnable.read_signals):
            short_signal = _short(signal_name)
            signal = signals_by_name.get(short_signal)
            if signal:
                _add_signal_atomic_port_from_access(
                    rows,
                    ports_added,
                    swc_name,
                    short_signal,
                    "R",
                    signal,
                    contract,
                )
            rows["RunnableAccesses"].append(
                _runnable_access_row(swc_name, runnable.runnable_name, "DataRead", short_signal, "", short_signal)
            )
        for signal_name in _split_list(runnable.write_signals):
            short_signal = _short(signal_name)
            signal = signals_by_name.get(short_signal)
            if signal:
                _add_signal_atomic_port_from_access(
                    rows,
                    ports_added,
                    swc_name,
                    short_signal,
                    "P",
                    signal,
                    contract,
                )
            rows["RunnableAccesses"].append(
                _runnable_access_row(swc_name, runnable.runnable_name, "DataWrite", short_signal, "", short_signal)
            )

    _dedupe_rows(rows)
    return rows


def _add_signal_atomic_data_type_rows(
    rows: dict[str, list[dict[str, str]]],
    data_type,
    mapping_path: str,
) -> None:
    type_name = _short(data_type.type_name)
    if not type_name:
        return
    kind = (data_type.type_kind or "").strip().lower()
    if kind == "record":
        return

    base_type = _normalize_base_type(data_type.base_type)
    app_ref = f"/DataTypes/{type_name}"
    impl_ref = f"/AUTOSAR_Platform/ImplementationDataTypes/{base_type}"
    category = _data_type_compu_category(data_type)
    compu_ref = _data_type_compu_ref(type_name, category, base_type)
    constr_ref = _data_type_constr_ref(type_name, category, data_type)

    rows["PrimitiveDataTypes"].append(
        {
            "ApplicationTypeName": type_name,
            "ApplicationTypePath": app_ref,
            "ImplementationTypeName": base_type,
            "ImplementationTypePath": impl_ref,
            "BaseType": base_type,
            "CompuMethodRef": compu_ref,
            "DataConstrRef": constr_ref,
            "CalibrationAccess": "READ-ONLY",
            "UnitRef": _signal_atomic_unit_ref(data_type.unit),
            "Description": data_type.description,
        }
    )
    rows["DataTypeMappings"].append(
        {
            "MappingSetPath": mapping_path,
            "ApplicationTypeRef": app_ref,
            "ImplementationTypeRef": impl_ref,
        }
    )

    if category == "IDENTICAL" and type_name == f"App_{base_type}":
        return

    compu_name = _short(compu_ref.rsplit("/", 1)[-1])
    rows["CompuMethods"].append(
        {
            "CompuMethodName": compu_name,
            "CompuMethodPath": compu_ref,
            "Category": category,
            "Description": data_type.description,
        }
    )
    if category == "TEXTTABLE":
        enum_values = data_type.enum_values or _boolean_enum_values(base_type)
        for lower, upper, enum_text in _split_enum_scales(enum_values):
            rows["CompuScales"].append(
                {
                    "CompuMethodName": compu_name,
                    "LowerLimit": lower,
                    "UpperLimit": upper,
                    "TextValue": enum_text,
                }
            )
    elif category == "LINEAR":
        lower, upper = _split_range(data_type.physical_range or data_type.internal_range)
        rows["CompuScales"].append(
            {
                "CompuMethodName": compu_name,
                "LowerLimit": lower,
                "UpperLimit": upper,
                "Numerator": data_type.resolution or "1",
                "Denominator": "1",
                "Offset": data_type.offset or "0",
            }
        )

    if constr_ref:
        lower, upper = _data_type_constr_range(data_type, category, base_type)
        rows["DataConstrs"].append(
            {
                "DataConstrName": _short(constr_ref.rsplit("/", 1)[-1]),
                "DataConstrPath": constr_ref,
                "LowerLimit": lower,
                "UpperLimit": upper,
                "Description": data_type.description,
            }
        )


def _signal_atomic_record_types(contract: DeliveryContract) -> list[dict[str, str]]:
    by_name: dict[str, dict[str, str]] = {}
    for signal in contract.signals:
        if (signal.value_type or "").lower() == "record" and signal.data_type:
            app = _short(signal.data_type)
            impl = _short(signal.internal_data_type) or f"Impl_{app.removeprefix('App_')}"
            by_name[app] = {"app": app, "impl": impl}
    for element in contract.record_elements:
        if element.record_type:
            app = _short(element.record_type)
            impl = _short(element.implementation_record_type) or f"Impl_{app.removeprefix('App_')}"
            by_name.setdefault(app, {"app": app, "impl": impl})
    return list(by_name.values())


def _add_signal_atomic_port_from_access(
    rows: dict[str, list[dict[str, str]]],
    ports_added: set[tuple[str, str]],
    swc_name: str,
    signal_name: str,
    direction: str,
    signal,
    contract: DeliveryContract,
) -> None:
    key = (swc_name, signal_name)
    if key in ports_added:
        return
    base_type = _normalize_base_type(getattr(signal, "internal_data_type", "") or signal.data_type)
    port_row = _sr_port_row(
        swc_name,
        signal_name,
        direction,
        f"/PortInterfaces/{signal_name}",
        signal_name,
        signal,
    )
    port_row["InitValue"] = _signal_init_value(signal, base_type)
    port_row["InitValueType"] = _signal_init_value_type(signal, base_type)
    rows["Ports"].append(port_row)
    if port_row["InitValueType"] == "Record":
        _add_record_init_rows(rows, swc_name, signal_name, signal, contract)
    ports_added.add(key)


def _add_record_init_rows(
    rows: dict[str, list[dict[str, str]]],
    swc_name: str,
    port_name: str,
    signal,
    contract: DeliveryContract,
) -> None:
    record_type = _short(signal.data_type)
    for element in sorted(contract.record_elements, key=lambda item: int(item.field_order or 0)):
        if _short(element.record_type) != record_type:
            continue
        field_name = element.element_name
        rows["PortRecordInitValues"].append(
            {
                "ComponentName": swc_name,
                "PortName": port_name,
                "RecordElementPath": field_name,
                "Value": element.init_value or "0",
                "ValueType": element.field_category or "Value",
                "Description": f"Init for {record_type}.{field_name}",
            }
        )


def _operation_event_bindings(contract: DeliveryContract) -> dict[tuple[str, str], dict[str, str]]:
    bindings: dict[tuple[str, str], dict[str, str]] = {}
    for runnable in contract.runnables:
        if runnable.trigger_type:
            continue
        if not runnable.related_port_or_signal or not runnable.related_operation:
            continue
        key = (runnable.swc, runnable.runnable_name)
        bindings[key] = {
            "port": runnable.related_port_or_signal,
            "operation": runnable.related_operation,
        }
    return bindings


def _infer_operation_event_binding(contract: DeliveryContract, runnable: RunnableContract) -> dict[str, str]:
    server_services = [
        service
        for service in contract.services
        if service.owner_swc == runnable.swc
        and _port_direction_from_role(service.port_role) == "P"
        and service.port_name
        and service.operation_name
    ]
    if len(server_services) == 1:
        service = server_services[0]
        return {"port": service.port_name, "operation": service.operation_name}

    runnable_key = _short(runnable.runnable_name).lower()
    for service in server_services:
        operation_key = _short(service.operation_name).lower()
        port_key = _short(service.port_name).lower()
        if operation_key and operation_key in runnable_key:
            return {"port": service.port_name, "operation": service.operation_name}
        if port_key and port_key in runnable_key:
            return {"port": service.port_name, "operation": service.operation_name}
    return {}


def _sr_port_row(component: str, port: str, direction: str, interface_ref: str, data_element: str, signal) -> dict[str, str]:
    return {
        "ComponentName": component,
        "PortName": port,
        "PortDirection": direction,
        "InterfaceKind": "SR",
        "InterfaceRef": interface_ref,
        "DataElementName": data_element,
        "ComSpecKind": "NONQUEUED-SENDER-COM-SPEC" if direction == "P" else "NONQUEUED-RECEIVER-COM-SPEC",
        "AliveTimeout": "0" if direction == "R" else "",
        "EnableUpdate": "false" if direction == "R" else "",
        "HandleNeverReceived": "false" if direction == "R" else "",
        "HandleTimeoutType": "NONE" if direction == "R" else "",
        "InitValue": signal.init_value or "0",
        "InitValueType": _init_value_type(signal),
        "Description": signal.description,
    }


def _cs_port_row(component: str, port: str, direction: str, interface_ref: str, operation: str) -> dict[str, str]:
    return {
        "ComponentName": component,
        "PortName": port,
        "PortDirection": direction,
        "InterfaceKind": "CS",
        "InterfaceRef": interface_ref,
        "OperationName": operation,
        "ComSpecKind": "SERVER-COM-SPEC" if direction == "P" else "CLIENT-COM-SPEC",
        "QueueLength": "1" if direction == "P" else "",
        "AliveTimeout": "0" if direction == "R" else "",
    }


def _add_validations(workbook: Workbook) -> None:
    rules = {
        "Components": {"B": '"Application,Composition"'},
        "PrimitiveDataTypes": {"E": '"boolean,uint8,uint16,uint32,uint64,sint8,sint16,sint32,float32"', "H": '"READ-ONLY,READ-WRITE,NOT-ACCESSIBLE"'},
        "CompuMethods": {"C": '"TEXTTABLE,LINEAR,IDENTICAL"'},
        "SRInterfaces": {"C": '"true,false"'},
        "CSInterfaces": {"C": '"true,false"'},
        "CSArguments": {"D": '"IN,OUT,INOUT"'},
        "Ports": {"C": '"P,R"', "D": '"SR,CS"', "H": '"CLIENT-COM-SPEC,SERVER-COM-SPEC,NONQUEUED-SENDER-COM-SPEC,NONQUEUED-RECEIVER-COM-SPEC,QUEUED-SENDER-COM-SPEC,QUEUED-RECEIVER-COM-SPEC"', "L": '"true,false"', "O": '"Value,Numeric,Enum,Boolean,String,Record"'},
        "RunnableEvents": {"C": '"Init,Periodic,OperationInvoked,DataReceived"'},
        "RunnableAccesses": {"C": '"DataRead,DataWrite,ServerCallPoint"'},
        "CompositionConnectors": {"F": '"Assembly,Delegation"'},
    }
    for sheet_name, column_rules in rules.items():
        sheet = workbook[sheet_name]
        for column, formula in column_rules.items():
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            sheet.add_data_validation(validation)
            validation.add(f"{column}2:{column}1000")


def _dedupe_rows(rows: dict[str, list[dict[str, str]]]) -> None:
    for sheet, sheet_rows in rows.items():
        seen = set()
        unique = []
        for row in sheet_rows:
            key = _dedupe_key(sheet, row)
            if key in seen:
                continue
            seen.add(key)
            unique.append(row)
        rows[sheet] = unique


def _dedupe_key(sheet: str, row: dict[str, str]) -> tuple[object, ...]:
    keys_by_sheet = {
        "Components": ("ComponentName",),
        "ComponentPrototypes": ("CompositionName", "PrototypeName"),
        "PrimitiveDataTypes": ("ApplicationTypePath",),
        "RecordTypes": ("ApplicationTypePath",),
        "RecordElements": ("RecordTypeName", "ElementName"),
        "PortRecordInitValues": ("ComponentName", "PortName", "RecordElementPath"),
        "DataTypeMappings": ("MappingSetPath", "ApplicationTypeRef", "ImplementationTypeRef"),
        "CompuMethods": ("CompuMethodName",),
        "CompuScales": ("CompuMethodName", "LowerLimit", "UpperLimit", "TextValue"),
        "DataConstrs": ("DataConstrPath",),
        "SRInterfaces": ("InterfaceName",),
        "SRDataElements": ("InterfaceName", "DataElementName"),
        "CSInterfaces": ("InterfaceName",),
        "CSOperations": ("InterfaceName", "OperationName"),
        "CSArguments": ("InterfaceName", "OperationName", "ArgumentName"),
        "Ports": ("ComponentName", "PortName"),
        "Runnables": ("ComponentName", "RunnableName"),
        "RunnableEvents": ("ComponentName", "RunnableName", "TriggerType", "PortName", "OperationName", "DataElementName"),
        "RunnableAccesses": ("ComponentName", "RunnableName", "AccessType", "PortName", "OperationName", "DataElementName"),
        "CompositionConnectors": ("CompositionName", "ProviderPrototype", "ProviderPort", "RequesterPrototype", "RequesterPort"),
        "Units": ("UnitPath",),
    }
    keys = keys_by_sheet.get(sheet)
    if not keys:
        return tuple(sorted(row.items()))
    return tuple(row.get(key, "") for key in keys)


def _root(contract: DeliveryContract) -> str:
    return contract.project.root_package or "/ARXML_PROJECT"


def _profile(contract: DeliveryContract) -> str:
    return (contract.project.generation_profile or contract.metadata.get("generation_profile") or "generic").strip()


def _atomic_swc_name(contract: DeliveryContract) -> str:
    for swc in contract.swcs:
        if (swc.kind or "Application").lower() == "application" and swc.name:
            return _short(swc.name)
    return _short(contract.project.system_name) or "AtomicSwc"


def _step_runnable_name(contract: DeliveryContract, swc_name: str) -> str:
    for runnable in contract.runnables:
        if runnable.runnable_name and runnable.trigger_type == "Periodic":
            return runnable.runnable_name
    for runnable in contract.runnables:
        if runnable.runnable_name and runnable.swc == swc_name:
            return runnable.runnable_name
    return ""


def _base_type_for_signal(contract: DeliveryContract, signal) -> str:
    wanted = _short(signal.data_type)
    for data_type in contract.data_types:
        if _short(data_type.type_name) == wanted:
            return _normalize_base_type(data_type.base_type)
    return _normalize_base_type(getattr(signal, "internal_data_type", "") or signal.data_type)


def _signal_has_texttable(signal) -> bool:
    base = _normalize_base_type(signal.data_type)
    return _is_enum_signal(signal) and bool((signal.enum_values or "").strip()) and base != "boolean"


def _signal_port_direction(signal, swc_name: str) -> str:
    direction = (signal.direction or "").strip().lower()
    if direction in {"output", "out", "p", "sender"}:
        return "P"
    if direction in {"input", "in", "r", "receiver"}:
        return "R"
    if signal.provider_swc == swc_name:
        return "P"
    return "R"


def _signal_app_type_ref(signal, base_type: str) -> str:
    if _short(signal.data_type):
        return f"/DataTypes/{_short(signal.data_type)}"
    return f"/DataTypes/App_{base_type}"


def _signal_init_value(signal, base_type: str) -> str:
    if _signal_has_texttable(signal):
        return _enum_init_value(signal.init_value, signal.enum_values)
    return "0" if base_type != "boolean" else _normalize_boolean_init(signal.init_value)


def _signal_init_value_type(signal, base_type: str) -> str:
    if (getattr(signal, "value_type", "") or "").strip().lower() == "record":
        return "Record"
    if _signal_has_texttable(signal):
        return "Enum"
    return "Boolean" if base_type == "boolean" else "Value"


def _normalize_boolean_init(value: str) -> str:
    return "1" if str(value or "").strip().lower() in {"1", "true", "on", "yes"} else "0"


def _is_enum_signal(signal) -> bool:
    value_type = (getattr(signal, "value_type", "") or "").strip().lower()
    return value_type == "enum"


def _data_type_compu_category(data_type) -> str:
    category = (getattr(data_type, "compu_method_category", "") or "").strip().upper()
    if category in {"TEXTTABLE", "LINEAR", "IDENTICAL"}:
        return category
    kind = (getattr(data_type, "type_kind", "") or "").strip().lower()
    if kind in {"enum", "boolean"}:
        return "TEXTTABLE"
    if _has_linear_conversion(getattr(data_type, "resolution", ""), getattr(data_type, "offset", "")):
        return "LINEAR"
    return "IDENTICAL"


def _data_type_compu_ref(type_name: str, category: str, base_type: str) -> str:
    if type_name == f"App_{base_type}" and category == "IDENTICAL":
        return _base_type_compu_method_ref(base_type)
    if base_type == "boolean" and category == "TEXTTABLE":
        return "/AUTOSAR_Platform/CompuMethods/boolean_CompuMethod"
    suffix = {"TEXTTABLE": "TextTable", "LINEAR": "Linear", "IDENTICAL": "Identical"}.get(category, "Identical")
    return f"/DataTypes/CompuMethods/CM_{type_name}_{suffix}"


def _data_type_constr_ref(type_name: str, category: str, data_type) -> str:
    base_type = _normalize_base_type(data_type.base_type)
    if type_name == f"App_{base_type}" and category == "IDENTICAL":
        return f"/AUTOSAR_Platform/DataConstrs/{base_type}_DataConstr"
    if base_type == "boolean":
        return "/AUTOSAR_Platform/DataConstrs/boolean_DataConstr"
    return f"/DataTypes/DataConstrs/DC_{type_name}"


def _data_type_constr_range(data_type, category: str, base_type: str) -> tuple[str, str]:
    if data_type.internal_range:
        return _split_range(data_type.internal_range)
    if category == "TEXTTABLE" and base_type == "boolean":
        return "0", "1"
    return "0", _base_type_upper(base_type)


def _boolean_enum_values(base_type: str) -> str:
    return "0=false;1=true" if base_type == "boolean" else ""


def _has_linear_conversion(resolution: str = "", offset: str = "") -> bool:
    resolution_text = str(resolution or "").strip()
    offset_text = str(offset or "").strip()
    if resolution_text and resolution_text not in {"1", "1.0"}:
        return True
    if offset_text and offset_text not in {"0", "0.0"}:
        return True
    return False


def _enum_init_value(init_value: str, enum_values: str) -> str:
    value = (init_value or "").strip()
    scales = _split_enum_scales(enum_values)
    if not scales:
        return "VALUE_0"
    if not value or value == "-":
        return scales[0][2]
    if re.match(r"^(0x[0-9A-Fa-f]+|\d+)$", value):
        parsed_value = _parse_int(value)
        for lower, upper, label in scales:
            try:
                if int(lower) <= int(parsed_value) <= int(upper):
                    return label
            except ValueError:
                continue
    return _short(value)


def _normalize_base_type(value: str) -> str:
    value = (value or "").strip().lower()
    if value in {"boolean", "bool"}:
        return "boolean"
    if value in {"uint8", "uint16", "uint32", "uint64", "sint8", "sint16", "sint32", "float32"}:
        return value
    return "uint8"


def _base_type_upper(base_type: str) -> str:
    return {
        "boolean": "1",
        "uint8": "255",
        "uint16": "65535",
        "uint32": "4294967295",
        "uint64": "18446744073709551615",
        "sint8": "127",
        "sint16": "32767",
        "sint32": "2147483647",
        "float32": "3.4028235e38",
    }.get(base_type, "255")


def _base_type_compu_method_name(base_type: str) -> str:
    return "boolean_CompuMethod" if base_type == "boolean" else f"CM_App_{base_type}_Identical"


def _base_type_compu_method_ref(base_type: str) -> str:
    if base_type == "boolean":
        return "/AUTOSAR_Platform/CompuMethods/boolean_CompuMethod"
    return f"/DataTypes/CompuMethods/{_base_type_compu_method_name(base_type)}"


def _clean_package(value: str, fallback: str) -> str:
    value = (value or "").strip() or fallback
    return "/" + value.strip("/")


def _davinci_paths(contract: DeliveryContract) -> dict[str, str]:
    project = contract.project
    return {
        "interface_package": _clean_package(project.interface_package, "/PortInterfaces"),
        "data_type_package": _clean_package(project.data_type_package, "/DataTypes"),
        "compu_method_package": _clean_package(project.compu_method_package, "/DataTypes/CompuMethods"),
        "data_constr_package": _clean_package(project.data_constr_package, "/DataTypes/DataConstrs"),
        "unit_package": _clean_package(project.unit_package, "/DataTypes/Units"),
        "mapping_set_path": _clean_package(project.mapping_set_path, "/ComponentTypes/MappingSets/APP_data_mapping"),
    }


def _interface_ref(paths: dict[str, str], name: str) -> str:
    return f'{paths["interface_package"]}/{_short(name)}'


def _data_type_ref(paths: dict[str, str], name: str) -> str:
    return f'{paths["data_type_package"]}/{_short(name)}'


def _app_type(name: str) -> str:
    short = _short(name)
    return short if short.startswith("App_") else f"App_{short}"


def _unit_ref(paths: dict[str, str], unit: str) -> str:
    return f'{paths["unit_package"]}/Unit_{_short(unit)}' if unit else ""


def _signal_atomic_unit_ref(unit: str) -> str:
    short = _short(unit)
    if not short:
        return ""
    if short.lower() == "no_unit":
        return "/DataTypes/Units/No_Unit"
    return f"/DataTypes/Units/Unit_{short}"


def _shared_base_type_unit_ref(contract: DeliveryContract, base_type: str) -> str:
    app_type = f"App_{base_type}"
    explicit_units = {
        (data_type.unit or "").strip()
        for data_type in contract.data_types
        if _short(data_type.type_name) == app_type and (data_type.unit or "").strip()
    }
    if not explicit_units:
        return ""
    return _signal_atomic_unit_ref(sorted(explicit_units)[0])



def _short(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    value = re.sub(r"[^A-Za-z0-9_]+", "_", value).strip("_")
    if re.match(r"^\d", value):
        value = "N_" + value
    return value


def _split_enum_values(value: str) -> list[str]:
    parts = re.split(r"[,，;；/、\n]+", value or "")
    cleaned = []
    for part in parts:
        text = part.strip()
        if not text:
            continue
        if "=" in text:
            text = text.split("=", 1)[-1].strip()
        cleaned.append(_short(text))
    return cleaned or ["VALUE_0"]


def _split_range(value: str) -> tuple[str, str]:
    match = re.search(r"(-?\d+(?:\.\d+)?)\s*[-~/至]\s*(-?\d+(?:\.\d+)?)", value or "")
    if match:
        return match.group(1), match.group(2)
    return "0", "255"


def _find_interface_for_operation(
    contract: DeliveryContract,
    operation: str,
    explicit_interface: str = "",
) -> str:
    if _short(explicit_interface):
        return _short(explicit_interface)
    for service in contract.services:
        if service.operation_name == operation:
            return service.interface_name or f"If_{service.service_name}_CS"
    return ""


def _is_client_service_port(
    contract: DeliveryContract,
    component_name: str,
    port_name: str,
    operation_name: str,
) -> bool:
    component = _short(component_name)
    port = _short(port_name)
    operation = _short(operation_name)
    for service in contract.services:
        if (
            _short(service.owner_swc or service.client_swc) == component
            and _short(service.port_name) == port
            and _short(service.operation_name) == operation
        ):
            return _port_direction_from_role(service.port_role) == "R"
    return False


def _init_value_type(signal) -> str:
    data_type = (signal.data_type or "").lower()
    if data_type == "boolean" or data_type.endswith("_boolean") or data_type.startswith("bool"):
        return "Boolean"
    if signal.enum_values:
        return "Enum"
    return "Numeric"


def _infer_operation_port(contract: DeliveryContract, runnable) -> str:
    operation = _short(runnable.related_operation)
    if not operation:
        return ""
    for service in contract.services:
        if service.owner_swc == runnable.swc and _short(service.operation_name) == operation:
            return _short(service.port_name)
        if service.provider_swc == runnable.swc and _short(service.operation_name) == operation:
            return _short(service.port_name) or f"Pp_{_short(service.service_name)}"
    for service in contract.services:
        if service.client_swc == runnable.swc and _short(service.operation_name) == operation:
            return _short(service.port_name) or f"Rp_{_short(service.service_name)}"
    return ""


def _composition_name(contract: DeliveryContract) -> str:
    if _short(contract.project.composition_name):
        return _short(contract.project.composition_name)
    for swc in contract.swcs:
        if (swc.kind or "").lower() == "composition":
            return swc.name
    system = _short(contract.project.system_name)
    return f"Composition_{system}" if system else "Composition_ARXML_Project"


def _prototype_name(component_name: str) -> str:
    short = _short(component_name)
    return f"{short}_Inst"


def _prototype_for_component(contract: DeliveryContract, component_name: str) -> str:
    short = _short(component_name)
    for swc in contract.swcs:
        if _short(swc.name) == short and _short(swc.prototype_name):
            return _short(swc.prototype_name)
    return _prototype_name(short)


def _connector_row(
    contract: DeliveryContract,
    composition_name: str,
    provider_component: str,
    provider_port: str,
    requester_component: str,
    requester_port: str,
    description: str,
) -> dict[str, str]:
    return {
        "CompositionName": composition_name,
        "ProviderPrototype": _prototype_for_component(contract, provider_component),
        "ProviderPort": provider_port,
        "RequesterPrototype": _prototype_for_component(contract, requester_component),
        "RequesterPort": requester_port,
        "ConnectorType": "Assembly",
        "Description": description,
    }


def _connector_row_from_contract(contract: DeliveryContract, composition_name: str, connector) -> dict[str, str]:
    provider_owner, provider_port = _split_endpoint(connector.provider_endpoint)
    requester_owner, requester_port = _split_endpoint(connector.requester_endpoint)
    if not provider_owner or not provider_port or not requester_owner or not requester_port:
        return {}
    return {
        "CompositionName": composition_name,
        "ProviderPrototype": provider_owner,
        "ProviderPort": provider_port,
        "RequesterPrototype": requester_owner,
        "RequesterPort": requester_port,
        "ConnectorType": connector.connector_type or "Assembly",
        "Description": connector.description,
    }


def _split_endpoint(endpoint: str) -> tuple[str, str]:
    text = (endpoint or "").strip()
    if "." not in text:
        return "", ""
    owner, port = text.split(".", 1)
    return _short(owner), _short(port)


def _port_direction_from_role(role: str) -> str:
    text = (role or "").strip().lower()
    if text in {"server", "sender", "provider", "p"}:
        return "P"
    if text in {"client", "receiver", "requester", "r"}:
        return "R"
    return ""


def _access_port_name(value: str, direction: str) -> str:
    short = _short(value)
    if short.startswith(("Rp_", "Pp_")):
        return short
    return f"{'Rp' if direction == 'R' else 'Pp'}_{short}"


def _data_element_from_port_or_signal(value: str) -> str:
    short = _short(value)
    if short.startswith(("Rp_", "Pp_")):
        return ""
    return f"DE_{short}"


def _runnable_access_row(
    component: str,
    runnable: str,
    access_type: str,
    port: str,
    operation: str,
    data_element: str,
) -> dict[str, str]:
    target = operation or data_element
    return {
        "ComponentName": component,
        "RunnableName": runnable,
        "AccessType": access_type,
        "PortName": port,
        "OperationName": operation,
        "DataElementName": data_element,
        "AccessName": f"{access_type}_{_short(port)}_{_short(target)}",
    }


def _split_list(value: str) -> list[str]:
    return [
        item.strip()
        for item in re.split(r"[,，;；、\n]+", value or "")
        if item.strip() and item.strip() != "-"
    ]


def _base_type_for_contract(contract: DeliveryContract, type_name: str) -> str:
    normalized = _short(type_name)
    for data_type in contract.data_types:
        if _short(data_type.type_name) == normalized:
            return data_type.base_type or "uint8"
    return "uint8"


def _related_port(value: str) -> str:
    return value if value.startswith(("Rp_", "Pp_")) else ""


def _related_data_element(value: str) -> str:
    return f"DE_{value}" if value and not value.startswith(("Rp_", "Pp_")) else ""


def _split_enum_scales(value: str) -> list[tuple[str, str, str]]:
    parts = re.split(r"[,，;；\n]+", value or "")
    scales: list[tuple[str, str, str]] = []
    next_value = 0
    for part in parts:
        text = part.strip()
        if not text:
            continue
        lower = upper = str(next_value)
        label = text
        if "=" in text:
            raw, label = text.split("=", 1)
            lower, upper = _enum_limit_pair(raw.strip())
            next_value = int(upper) + 1 if upper.isdigit() else next_value + 1
        else:
            next_value += 1
        scales.append((lower, upper, _short(label.strip())))
    return scales or [("0", "0", "VALUE_0")]


def _enum_limit_pair(value: str) -> tuple[str, str]:
    match = re.match(r"^(0x[0-9A-Fa-f]+|\d+)\s*[-~]\s*(0x[0-9A-Fa-f]+|\d+)$", value)
    if match:
        return _parse_int(match.group(1)), _parse_int(match.group(2))
    parsed = _parse_int(value)
    return parsed, parsed


def _parse_int(value: str) -> str:
    value = value.strip()
    return str(int(value, 16)) if value.lower().startswith("0x") else value
