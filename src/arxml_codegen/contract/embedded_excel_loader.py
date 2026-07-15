from __future__ import annotations

import re
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from arxml_codegen.contract.schema import (
    DataTypeContract,
    DeliveryContract,
    OpenIssue,
    ProjectContract,
    RunnableContract,
    SignalContract,
    SwcContract,
)


def extract_signal_atomic_contract_from_docx(path: Path) -> DeliveryContract:
    """Extract a signal-atomic contract from a DOCX with embedded Excel sheets.

    Some upstream delivery documents look like Word tables to humans but store
    the editable content as embedded XLSX objects. This loader treats the first
    embedded workbook as runnable config, the second as input signals, and the
    third as output signals when headers cannot be recognized reliably.
    """

    embedded = _extract_embedded_workbooks(path)
    role_rows = _classify_workbooks(embedded)
    runnable_rows = role_rows.get("runnables", [])
    input_rows = role_rows.get("inputs", [])
    output_rows = role_rows.get("outputs", [])

    swc_name = _guess_swc_name(path, runnable_rows, output_rows) or "AtomicSwc"
    contract = DeliveryContract(
        project=ProjectContract(
            system_name=swc_name,
            root_package="/ComponentTypes",
            generation_profile="signal_atomic_davinci",
            mapping_set_path="/ComponentTypes/MappingSets/DataMapping",
            source_status={
                "system_name": "inferred",
                "root_package": "defaulted",
                "generation_profile": "defaulted",
                "target_autosar_version": "defaulted",
            },
        ),
        swcs=[
            SwcContract(
                name=swc_name,
                kind="Application",
                description="Single atomic SWC from embedded Excel signal delivery document",
            )
        ],
        metadata={
            "source_docx": str(path),
            "mode": "signal",
            "generation_profile": "signal_atomic_davinci",
            "source_shape": "docx_with_embedded_xlsx",
        },
    )

    for row in runnable_rows:
        swc = _short(_at(row, 0)) or swc_name
        runnable = _short(_at(row, 1))
        if not runnable:
            continue
        trigger = _at(row, 2).lower()
        period = _at(row, 3).lower().replace("ms", "").strip()
        if period == "-":
            period = ""
        contract.runnables.append(
            RunnableContract(
                swc=swc,
                runnable_name=runnable,
                trigger_type="Periodic" if "period" in trigger else "Init",
                period_ms=period,
                description=_at(row, 4),
                source_status={"runnable_name": "explicit", "trigger_type": "explicit"},
            )
        )

    seen_types: set[tuple[str, str, str]] = set()
    for row in input_rows:
        _append_signal(contract, row, swc_name=swc_name, is_input=True, seen_types=seen_types)
    for row in output_rows:
        _append_signal(contract, row, swc_name=swc_name, is_input=False, seen_types=seen_types)

    contract.open_issues.extend(
        [
            OpenIssue(
                field="ExternalBoundary",
                question="Signal atomic profile does not generate Composition/Connector; ports are treated as ECU/SWC boundary signals.",
                suggested_default="No CompositionConnector",
            ),
            OpenIssue(
                field="InitValue",
                question="InitValue is not a dedicated column in the source document. Numeric/boolean defaults to 0; enum defaults to the first text-table symbol.",
                suggested_default="0 or first enum symbol",
            ),
        ]
    )
    return contract


def _classify_workbooks(paths: list[Path]) -> dict[str, list[list[str]]]:
    classified: dict[str, list[list[str]]] = {}
    unclassified: list[list[list[str]]] = []
    for workbook_path in paths:
        headers, rows = _read_first_sheet(workbook_path)
        header_text = "|".join(headers).lower()
        if "runnable" in header_text or "runable" in header_text:
            classified["runnables"] = rows
        elif "input" in header_text or "输入" in header_text:
            classified["inputs"] = rows
        elif "output" in header_text or "输出" in header_text:
            classified["outputs"] = rows
        else:
            unclassified.append(rows)

    # WPS/Office embedded objects are often named generically. Use the standard
    # delivery order as fallback: runnable, inputs, outputs.
    ordered = [rows for _, rows in zip(paths, [_read_first_sheet(path)[1] for path in paths])]
    if "runnables" not in classified and len(ordered) >= 1:
        classified["runnables"] = ordered[0]
    if "inputs" not in classified and len(ordered) >= 2:
        classified["inputs"] = ordered[1]
    if "outputs" not in classified and len(ordered) >= 3:
        classified["outputs"] = ordered[2]
    return classified


def _extract_embedded_workbooks(path: Path) -> list[Path]:
    out_dir = Path("output") / f"{_short(path.stem)}_embedded"
    out_dir.mkdir(parents=True, exist_ok=True)
    extracted: list[Path] = []
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            if name.startswith("word/embeddings/") and name.endswith(".xlsx"):
                target = out_dir / Path(name).name
                target.write_bytes(archive.read(name))
                extracted.append(target)
    return sorted(extracted)


def _read_first_sheet(path: Path) -> tuple[list[str], list[list[str]]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.worksheets[0]
    rows = [[_clean(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    if not rows:
        return [], []
    return rows[0], [row for row in rows[1:] if any(row)]


def _append_signal(
    contract: DeliveryContract,
    row: list[str],
    *,
    swc_name: str,
    is_input: bool,
    seen_types: set[tuple[str, str, str]],
) -> None:
    signal_name = _short(_at(row, 1))
    if not signal_name:
        return
    base_type = _base_type(_at(row, 3))
    enum_values = "" if base_type == "boolean" else ", ".join(_enum_values(_at(row, 6)))
    type_name = signal_name if enum_values else base_type
    contract.signals.append(
        SignalContract(
            signal_name=signal_name,
            direction="R" if is_input else "P",
            provider_swc="" if is_input else swc_name,
            consumer_swc=swc_name if is_input else "",
            data_type=type_name,
            enum_values=enum_values,
            init_value=_first_enum(enum_values) if enum_values else "0",
            description=_description(row),
            source=_at(row, 7),
            source_status={
                "signal_name": "explicit",
                "data_type": "explicit",
                "enum_values": "explicit" if enum_values else "missing",
                "init_value": "inferred" if enum_values else "defaulted",
            },
        )
    )
    key = (type_name, base_type, enum_values)
    if key not in seen_types:
        seen_types.add(key)
        contract.data_types.append(
            DataTypeContract(
                type_name=type_name,
                base_type=base_type,
                compu_method_category="TEXTTABLE" if enum_values or base_type == "boolean" else "IDENTICAL",
                enum_values=enum_values,
                physical_range="0..1" if base_type == "boolean" else "",
                description="Derived from signal atomic delivery document",
                source_status={"type_name": "inferred", "base_type": "explicit"},
            )
        )


def _enum_values(detail: str) -> list[str]:
    values: list[str] = []
    for match in re.finditer(r'(0x[0-9A-Fa-f]+|\d+)\s*(?:[:：]|\s+)\s*"?([^"/\n]+)"?', detail or ""):
        raw_value = match.group(1)
        number = str(int(raw_value, 16)) if raw_value.lower().startswith("0x") else raw_value
        symbol = _short(match.group(2))
        item = f"{number}={symbol}" if symbol else ""
        if item and item not in values:
            values.append(item)
    return values


def _first_enum(enum_values: str) -> str:
    return next((item.strip() for item in enum_values.split(",") if item.strip()), "")


def _description(row: list[str]) -> str:
    labels = ["No", "Signal", "Module", "DataType", "Channel", "Description", "Detail", "CANSignal", "Scenario"]
    parts = []
    for index, label in enumerate(labels):
        value = _at(row, index)
        if value:
            parts.append(f"{label}: {value}")
    return "; ".join(parts)


def _guess_swc_name(path: Path, runnable_rows: list[list[str]], output_rows: list[list[str]]) -> str:
    for row in runnable_rows:
        if _at(row, 0):
            return _short(_at(row, 0))
    for row in output_rows:
        if _at(row, 2):
            return _short(_at(row, 2))
    match = re.search(r"([A-Za-z][A-Za-z0-9_]*)", path.stem)
    return _short(match.group(1)) if match else ""


def _base_type(value: str) -> str:
    value = value.strip().lower()
    if value in {"boolean", "bool"}:
        return "boolean"
    if value in {"uint8", "uint16", "uint32", "uint64", "sint8", "sint16", "sint32", "float32"}:
        return value
    return "uint8"


def _at(row: list[str], index: int) -> str:
    return row[index] if index < len(row) else ""


def _clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def _short(value: str) -> str:
    value = (value or "").strip()
    value = re.sub(r"[^A-Za-z0-9_]+", "_", value).strip("_")
    if value and value[0].isdigit():
        value = "N_" + value
    return value
