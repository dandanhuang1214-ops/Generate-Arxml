from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from arxml_codegen.models.schema import (
    CSArgumentRow,
    CSInterfaceRow,
    CSOperationRow,
    CompuMethodRow,
    CompuScaleRow,
    ComponentPrototypeRow,
    ComponentV2Row,
    CompositionConnectorV2Row,
    DataConstrRow,
    DataTypeMappingRow,
    PrimitiveDataTypeRow,
    PortRecordInitValueRow,
    ProjectConfigRow,
    RecordElementRow,
    RecordTypeRow,
    RunnableAccessRow,
    RunnableEventV2Row,
    RunnableV2Row,
    SRDataElementRow,
    SRInterfaceRow,
    UnitRow,
    PortV2Row,
    WorkbookV2Model,
)


def _normalize(value: object) -> str:
    return "" if value is None else str(value).strip()


def _rows(workbook, sheet_name: str):
    if sheet_name not in workbook.sheetnames:
        return
    sheet = workbook[sheet_name]
    rows = [[_normalize(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    if not rows:
        return
    header = {name: idx for idx, name in enumerate(rows[0]) if name}
    for row_index, row in enumerate(rows[1:], start=2):
        if any(row):
            yield sheet_name, row_index, header, row


def _cell(row: list[str], header: dict[str, int], key: str) -> str:
    idx = header.get(key)
    if idx is None or idx >= len(row):
        return ""
    return row[idx]


def load_workbook_v2(path: Path) -> WorkbookV2Model:
    workbook = load_workbook(path, data_only=True)
    model = WorkbookV2Model()

    for sheet, row_index, header, row in _rows(workbook, "ProjectConfig") or []:
        model.project_config.append(ProjectConfigRow(sheet, row_index, _cell(row, header, "Key"), _cell(row, header, "Value")))

    for sheet, row_index, header, row in _rows(workbook, "Components") or []:
        model.components.append(
            ComponentV2Row(
                source_sheet=sheet,
                row_index=row_index,
                component_name=_cell(row, header, "ComponentName"),
                component_kind=_cell(row, header, "ComponentKind") or "Application",
                package_path=_cell(row, header, "PackagePath"),
                internal_behavior_name=_cell(row, header, "InternalBehaviorName"),
                implementation_name=_cell(row, header, "ImplementationName"),
            )
        )

    for sheet, row_index, header, row in _rows(workbook, "ComponentPrototypes") or []:
        model.component_prototypes.append(
            ComponentPrototypeRow(sheet, row_index, _cell(row, header, "CompositionName"), _cell(row, header, "PrototypeName"), _cell(row, header, "ComponentTypeName"), _cell(row, header, "ComponentTypeRef"))
        )

    for sheet, row_index, header, row in _rows(workbook, "PrimitiveDataTypes") or []:
        model.primitive_data_types.append(
            PrimitiveDataTypeRow(
                source_sheet=sheet,
                row_index=row_index,
                application_type_name=_cell(row, header, "ApplicationTypeName"),
                application_type_path=_cell(row, header, "ApplicationTypePath"),
                implementation_type_name=_cell(row, header, "ImplementationTypeName"),
                implementation_type_path=_cell(row, header, "ImplementationTypePath"),
                base_type=_cell(row, header, "BaseType"),
                compu_method_ref=_cell(row, header, "CompuMethodRef"),
                data_constr_ref=_cell(row, header, "DataConstrRef"),
                calibration_access=_cell(row, header, "CalibrationAccess") or "READ-ONLY",
                unit_ref=_cell(row, header, "UnitRef"),
            )
        )

    for sheet, row_index, header, row in _rows(workbook, "RecordTypes") or []:
        model.record_types.append(
            RecordTypeRow(sheet, row_index, _cell(row, header, "ApplicationTypeName"), _cell(row, header, "ApplicationTypePath"), _cell(row, header, "ImplementationTypeName"), _cell(row, header, "ImplementationTypePath"), _cell(row, header, "CalibrationAccess") or "READ-ONLY")
        )

    for sheet, row_index, header, row in _rows(workbook, "RecordElements") or []:
        model.record_elements.append(
            RecordElementRow(sheet, row_index, _cell(row, header, "RecordTypeName"), _cell(row, header, "ElementName"), _cell(row, header, "ApplicationElementTypeRef"), _cell(row, header, "ImplementationElementTypeRef"), _cell(row, header, "Order"))
        )

    for sheet, row_index, header, row in _rows(workbook, "PortRecordInitValues") or []:
        model.port_record_init_values.append(
            PortRecordInitValueRow(
                sheet,
                row_index,
                _cell(row, header, "ComponentName"),
                _cell(row, header, "PortName"),
                _cell(row, header, "RecordElementPath"),
                _cell(row, header, "Value"),
                _cell(row, header, "ValueType"),
            )
        )

    for sheet, row_index, header, row in _rows(workbook, "DataTypeMappings") or []:
        model.data_type_mappings.append(
            DataTypeMappingRow(sheet, row_index, _cell(row, header, "MappingSetPath"), _cell(row, header, "ApplicationTypeRef"), _cell(row, header, "ImplementationTypeRef"))
        )

    for sheet, row_index, header, row in _rows(workbook, "CompuMethods") or []:
        model.compu_methods.append(CompuMethodRow(sheet, row_index, _cell(row, header, "CompuMethodName"), _cell(row, header, "CompuMethodPath"), _cell(row, header, "Category") or "IDENTICAL"))

    for sheet, row_index, header, row in _rows(workbook, "CompuScales") or []:
        model.compu_scales.append(CompuScaleRow(sheet, row_index, _cell(row, header, "CompuMethodName"), _cell(row, header, "LowerLimit"), _cell(row, header, "UpperLimit"), _cell(row, header, "TextValue"), _cell(row, header, "Numerator"), _cell(row, header, "Denominator"), _cell(row, header, "Offset")))

    for sheet, row_index, header, row in _rows(workbook, "DataConstrs") or []:
        model.data_constrs.append(DataConstrRow(sheet, row_index, _cell(row, header, "DataConstrName"), _cell(row, header, "DataConstrPath"), _cell(row, header, "LowerLimit"), _cell(row, header, "UpperLimit")))

    for sheet, row_index, header, row in _rows(workbook, "SRInterfaces") or []:
        model.sr_interfaces.append(SRInterfaceRow(sheet, row_index, _cell(row, header, "InterfaceName"), _cell(row, header, "InterfacePath"), _cell(row, header, "IsService") or "false"))

    for sheet, row_index, header, row in _rows(workbook, "SRDataElements") or []:
        model.sr_data_elements.append(SRDataElementRow(sheet, row_index, _cell(row, header, "InterfaceName"), _cell(row, header, "DataElementName"), _cell(row, header, "ApplicationTypeRef")))

    for sheet, row_index, header, row in _rows(workbook, "CSInterfaces") or []:
        model.cs_interfaces.append(CSInterfaceRow(sheet, row_index, _cell(row, header, "InterfaceName"), _cell(row, header, "InterfacePath"), _cell(row, header, "IsService") or "false"))

    for sheet, row_index, header, row in _rows(workbook, "CSOperations") or []:
        model.cs_operations.append(CSOperationRow(sheet, row_index, _cell(row, header, "InterfaceName"), _cell(row, header, "OperationName")))

    for sheet, row_index, header, row in _rows(workbook, "CSArguments") or []:
        model.cs_arguments.append(CSArgumentRow(sheet, row_index, _cell(row, header, "InterfaceName"), _cell(row, header, "OperationName"), _cell(row, header, "ArgumentName"), _cell(row, header, "Direction"), _cell(row, header, "ApplicationTypeRef")))

    for sheet, row_index, header, row in _rows(workbook, "Ports") or []:
        model.ports.append(
            PortV2Row(
                sheet,
                row_index,
                _cell(row, header, "ComponentName"),
                _cell(row, header, "PortName"),
                _cell(row, header, "PortDirection"),
                _cell(row, header, "InterfaceKind"),
                _cell(row, header, "InterfaceRef"),
                _cell(row, header, "DataElementName"),
                _cell(row, header, "OperationName"),
                _cell(row, header, "ComSpecKind"),
                _cell(row, header, "AliveTimeout"),
                _cell(row, header, "QueueLength"),
                _cell(row, header, "EnableUpdate"),
                _cell(row, header, "HandleNeverReceived"),
                _cell(row, header, "HandleTimeoutType"),
                _cell(row, header, "InitValue"),
                _cell(row, header, "InitValueType"),
            )
        )

    for sheet, row_index, header, row in _rows(workbook, "Runnables") or []:
        model.runnables.append(RunnableV2Row(sheet, row_index, _cell(row, header, "ComponentName"), _cell(row, header, "RunnableName"), _cell(row, header, "Symbol")))

    for sheet, row_index, header, row in _rows(workbook, "RunnableEvents") or []:
        model.runnable_events.append(RunnableEventV2Row(sheet, row_index, _cell(row, header, "ComponentName"), _cell(row, header, "RunnableName"), _cell(row, header, "TriggerType"), _cell(row, header, "PeriodMs"), _cell(row, header, "PortName"), _cell(row, header, "OperationName"), _cell(row, header, "DataElementName")))

    for sheet, row_index, header, row in _rows(workbook, "RunnableAccesses") or []:
        model.runnable_accesses.append(RunnableAccessRow(sheet, row_index, _cell(row, header, "ComponentName"), _cell(row, header, "RunnableName"), _cell(row, header, "AccessType"), _cell(row, header, "PortName"), _cell(row, header, "OperationName"), _cell(row, header, "DataElementName"), _cell(row, header, "AccessName")))

    for sheet, row_index, header, row in _rows(workbook, "CompositionConnectors") or []:
        model.composition_connectors.append(CompositionConnectorV2Row(sheet, row_index, _cell(row, header, "CompositionName"), _cell(row, header, "ProviderPrototype"), _cell(row, header, "ProviderPort"), _cell(row, header, "RequesterPrototype"), _cell(row, header, "RequesterPort"), _cell(row, header, "ConnectorType") or "Assembly"))

    for sheet, row_index, header, row in _rows(workbook, "Units") or []:
        model.units.append(UnitRow(sheet, row_index, _cell(row, header, "UnitName"), _cell(row, header, "UnitPath"), _cell(row, header, "DisplayName"), _cell(row, header, "FactorSIToUnit") or "1", _cell(row, header, "OffsetSIToUnit") or "0"))

    return model
