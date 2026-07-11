from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


SHEETS: dict[str, list[str]] = {
    "ProjectConfig": ["Key", "Value", "Description"],
    "Components": ["ComponentName", "ComponentKind", "PackagePath", "InternalBehaviorName", "ImplementationName", "Description"],
    "ComponentPrototypes": ["CompositionName", "PrototypeName", "ComponentTypeName", "ComponentTypeRef", "Description"],
    "PrimitiveDataTypes": ["ApplicationTypeName", "ApplicationTypePath", "ImplementationTypeName", "ImplementationTypePath", "BaseType", "CompuMethodRef", "DataConstrRef", "CalibrationAccess", "UnitRef", "Description"],
    "RecordTypes": ["ApplicationTypeName", "ApplicationTypePath", "ImplementationTypeName", "ImplementationTypePath", "CalibrationAccess", "Description"],
    "RecordElements": ["RecordTypeName", "ElementName", "ApplicationElementTypeRef", "ImplementationElementTypeRef", "Order", "Description"],
    "DataTypeMappings": ["MappingSetPath", "ApplicationTypeRef", "ImplementationTypeRef", "Description"],
    "CompuMethods": ["CompuMethodName", "CompuMethodPath", "Category", "Description"],
    "CompuScales": ["CompuMethodName", "LowerLimit", "UpperLimit", "TextValue", "Numerator", "Denominator", "Offset", "Description"],
    "DataConstrs": ["DataConstrName", "DataConstrPath", "LowerLimit", "UpperLimit", "Description"],
    "SRInterfaces": ["InterfaceName", "InterfacePath", "IsService", "Description"],
    "SRDataElements": ["InterfaceName", "DataElementName", "ApplicationTypeRef", "Description"],
    "CSInterfaces": ["InterfaceName", "InterfacePath", "IsService", "Description"],
    "CSOperations": ["InterfaceName", "OperationName", "Description"],
    "CSArguments": ["InterfaceName", "OperationName", "ArgumentName", "Direction", "ApplicationTypeRef", "Description"],
    "Ports": ["ComponentName", "PortName", "PortDirection", "InterfaceKind", "InterfaceRef", "DataElementName", "OperationName", "ComSpecKind", "AliveTimeout", "QueueLength", "EnableUpdate", "HandleTimeoutType", "InitValue", "Description"],
    "Runnables": ["ComponentName", "RunnableName", "Symbol", "Description"],
    "RunnableEvents": ["ComponentName", "RunnableName", "TriggerType", "PeriodMs", "PortName", "OperationName", "DataElementName", "Description"],
    "RunnableAccesses": ["ComponentName", "RunnableName", "AccessType", "PortName", "OperationName", "DataElementName", "AccessName", "Description"],
    "CompositionConnectors": ["CompositionName", "ProviderPrototype", "ProviderPort", "RequesterPrototype", "RequesterPort", "ConnectorType", "Description"],
    "Units": ["UnitName", "UnitPath", "DisplayName", "FactorSIToUnit", "OffsetSIToUnit", "Description"],
}


EXAMPLE_ROWS: dict[str, list[list[str]]] = {
    "ProjectConfig": [
        ["AutosarVersion", "4-3-0", "AUTOSAR schema version"],
        ["RootPackage", "/HORN_CTRL", "Root package for the project"],
        ["DefaultMappingSetPath", "/ComponentTypes/MappingSets/APP_data_mapping", "Default data type mapping set path"],
    ],
    "Components": [
        ["Composition_HornCtrl", "Composition", "/HORN_CTRL/System", "", "", "Top-level composition SWC"],
        ["BOD_Horn_Atm", "Application", "/HORN_CTRL/Components", "IB_BOD_Horn_Atm", "BOD_Horn_Atm_Impl", "Atomic layer: horn actuator"],
        ["BOD_HornCtrl_Enh", "Application", "/HORN_CTRL/Components", "IB_BOD_HornCtrl_Enh", "BOD_HornCtrl_Enh_Impl", "Enhancement layer: horn control"],
        ["BOD_ChildLeftBehindAlert_Scen", "Application", "/HORN_CTRL/Components", "IB_ChildLeftBehindAlert_Scen", "BOD_ChildLeftBehindAlert_Scen_Impl", "Scenario layer: child-left-behind alert"],
        ["BOD_SerchCar_Gen", "Application", "/HORN_CTRL/Components", "IB_BOD_SerchCar_Gen", "BOD_SerchCar_Gen_Impl", "Scenario layer: car search"],
    ],
    "ComponentPrototypes": [
        ["Composition_HornCtrl", "Atm_Inst", "BOD_Horn_Atm", "/HORN_CTRL/Components/BOD_Horn_Atm", ""],
        ["Composition_HornCtrl", "Enh_Inst", "BOD_HornCtrl_Enh", "/HORN_CTRL/Components/BOD_HornCtrl_Enh", ""],
        ["Composition_HornCtrl", "Scen_Inst", "BOD_ChildLeftBehindAlert_Scen", "/HORN_CTRL/Components/BOD_ChildLeftBehindAlert_Scen", ""],
        ["Composition_HornCtrl", "Gen_Inst", "BOD_SerchCar_Gen", "/HORN_CTRL/Components/BOD_SerchCar_Gen", ""],
    ],
    "PrimitiveDataTypes": [
        ["App_HornCmd", "/HORN_CTRL/ApplicationDataTypes/App_HornCmd", "uint8", "/AUTOSAR_Platform/ImplementationDataTypes/uint8", "uint8", "/HORN_CTRL/CompuMethods/CM_HornCmd_TextTable", "/HORN_CTRL/DataConstrs/DC_App_HornCmd", "READ-ONLY", "", "Horn command: OFF=0, ON=1"],
        ["App_ReturnCode", "/HORN_CTRL/ApplicationDataTypes/App_ReturnCode", "uint8", "/AUTOSAR_Platform/ImplementationDataTypes/uint8", "uint8", "", "", "READ-ONLY", "", "Operation return code"],
        ["App_HornActSts", "/HORN_CTRL/ApplicationDataTypes/App_HornActSts", "uint8", "/AUTOSAR_Platform/ImplementationDataTypes/uint8", "uint8", "", "/HORN_CTRL/DataConstrs/DC_App_HornActSts", "READ-ONLY", "", "Horn activation status"],
        ["App_Volt", "/HORN_CTRL/ApplicationDataTypes/App_Volt", "uint16", "/AUTOSAR_Platform/ImplementationDataTypes/uint16", "uint16", "/HORN_CTRL/CompuMethods/CM_Volt_Linear", "/HORN_CTRL/DataConstrs/DC_App_Volt", "READ-ONLY", "/HORN_CTRL/Units/Unit_Volt", "Vehicle voltage in 0.1V"],
        ["App_Curr", "/HORN_CTRL/ApplicationDataTypes/App_Curr", "uint16", "/AUTOSAR_Platform/ImplementationDataTypes/uint16", "uint16", "/HORN_CTRL/CompuMethods/CM_Curr_Linear", "/HORN_CTRL/DataConstrs/DC_App_Curr", "READ-ONLY", "/HORN_CTRL/Units/Unit_Ampere", "Horn current in 0.01A"],
        ["App_FltSts", "/HORN_CTRL/ApplicationDataTypes/App_FltSts", "uint8", "/AUTOSAR_Platform/ImplementationDataTypes/uint8", "uint8", "", "", "READ-ONLY", "", "Fault status"],
        ["App_IsEnable", "/HORN_CTRL/ApplicationDataTypes/App_IsEnable", "boolean", "/AUTOSAR_Platform/ImplementationDataTypes/boolean", "boolean", "", "", "READ-ONLY", "", "Boolean enable flag"],
        ["App_DutyRat", "/HORN_CTRL/ApplicationDataTypes/App_DutyRat", "uint16", "/AUTOSAR_Platform/ImplementationDataTypes/uint16", "uint16", "/HORN_CTRL/CompuMethods/CM_DutyRat_Linear", "/HORN_CTRL/DataConstrs/DC_App_DutyRat", "READ-ONLY", "", "Duty ratio in 0.1%"],
        ["App_HornOperCmd", "/HORN_CTRL/ApplicationDataTypes/App_HornOperCmd", "uint8", "/AUTOSAR_Platform/ImplementationDataTypes/uint8", "uint8", "/HORN_CTRL/CompuMethods/CM_HornOperCmd_TextTable", "", "READ-ONLY", "", "Horn operation command"],
        ["App_OnDuration", "/HORN_CTRL/ApplicationDataTypes/App_OnDuration", "uint32", "/AUTOSAR_Platform/ImplementationDataTypes/uint32", "uint32", "", "", "READ-ONLY", "", "ON duration in ms"],
        ["App_OffDuration", "/HORN_CTRL/ApplicationDataTypes/App_OffDuration", "uint32", "/AUTOSAR_Platform/ImplementationDataTypes/uint32", "uint32", "", "", "READ-ONLY", "", "OFF duration in ms"],
        ["App_HornTimes", "/HORN_CTRL/ApplicationDataTypes/App_HornTimes", "uint8", "/AUTOSAR_Platform/ImplementationDataTypes/uint8", "uint8", "/HORN_CTRL/CompuMethods/CM_HornTimes", "", "READ-ONLY", "", "Horn repeat times"],
        ["APP_Boolean", "/DataTypes/APP_Boolean", "boolean", "/AUTOSAR_Platform/ImplementationDataTypes/boolean", "boolean", "/AUTOSAR_Platform/CompuMethods/boolean_CompuMethod", "/AUTOSAR_Platform/DataConstrs/boolean_DataConstr", "READ-ONLY", "", "Platform boolean ADT"],
    ],
    "RecordTypes": [
        ["App_HornCtrlMode", "/HORN_CTRL/ApplicationDataTypes/App_HornCtrlMode", "Impl_HornCtrlMode", "/HORN_CTRL/ImplementationDataTypes/Impl_HornCtrlMode", "READ-ONLY", "CS operation argument record"],
        ["App_HornPeriodMode", "/HORN_CTRL/ApplicationDataTypes/App_HornPeriodMode", "Impl_HornPeriod", "/HORN_CTRL/ImplementationDataTypes/Impl_HornPeriod", "READ-ONLY", "Period config record"],
        ["App_SrvOperSts", "/HORN_CTRL/ApplicationDataTypes/App_SrvOperSts", "Impl_SrvOperSts_Struct", "/HORN_CTRL/ImplementationDataTypes/Impl_SrvOperSts_Struct", "READ-ONLY", "Service operation status record"],
    ],
    "RecordElements": [
        ["App_HornCtrlMode", "HornCmd", "/HORN_CTRL/ApplicationDataTypes/App_HornCmd", "/AUTOSAR_Platform/ImplementationDataTypes/uint8", "1", ""],
        ["App_HornCtrlMode", "Volt", "/HORN_CTRL/ApplicationDataTypes/App_Volt", "/AUTOSAR_Platform/ImplementationDataTypes/uint16", "2", ""],
        ["App_HornCtrlMode", "Enable", "/HORN_CTRL/ApplicationDataTypes/App_IsEnable", "/AUTOSAR_Platform/ImplementationDataTypes/boolean", "3", ""],
        ["App_HornPeriodMode", "OnDuration", "/HORN_CTRL/ApplicationDataTypes/App_OnDuration", "/AUTOSAR_Platform/ImplementationDataTypes/uint32", "1", ""],
        ["App_HornPeriodMode", "OffDuration", "/HORN_CTRL/ApplicationDataTypes/App_OffDuration", "/AUTOSAR_Platform/ImplementationDataTypes/uint32", "2", ""],
        ["App_HornPeriodMode", "HornTimes", "/HORN_CTRL/ApplicationDataTypes/App_HornTimes", "/AUTOSAR_Platform/ImplementationDataTypes/uint8", "3", ""],
    ],
    "DataTypeMappings": [
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_HornCmd", "/AUTOSAR_Platform/ImplementationDataTypes/uint8", ""],
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_ReturnCode", "/AUTOSAR_Platform/ImplementationDataTypes/uint8", ""],
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_HornActSts", "/AUTOSAR_Platform/ImplementationDataTypes/uint8", ""],
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_Volt", "/AUTOSAR_Platform/ImplementationDataTypes/uint16", ""],
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_Curr", "/AUTOSAR_Platform/ImplementationDataTypes/uint16", ""],
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_FltSts", "/AUTOSAR_Platform/ImplementationDataTypes/uint8", ""],
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_IsEnable", "/AUTOSAR_Platform/ImplementationDataTypes/boolean", ""],
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_DutyRat", "/AUTOSAR_Platform/ImplementationDataTypes/uint16", ""],
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_HornOperCmd", "/AUTOSAR_Platform/ImplementationDataTypes/uint8", ""],
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_OnDuration", "/AUTOSAR_Platform/ImplementationDataTypes/uint32", ""],
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_OffDuration", "/AUTOSAR_Platform/ImplementationDataTypes/uint32", ""],
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_HornTimes", "/AUTOSAR_Platform/ImplementationDataTypes/uint8", ""],
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_HornCtrlMode", "/HORN_CTRL/ImplementationDataTypes/Impl_HornCtrlMode", ""],
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_HornPeriodMode", "/HORN_CTRL/ImplementationDataTypes/Impl_HornPeriod", ""],
        ["/ComponentTypes/MappingSets/APP_data_mapping", "/HORN_CTRL/ApplicationDataTypes/App_SrvOperSts", "/HORN_CTRL/ImplementationDataTypes/Impl_SrvOperSts_Struct", ""],
    ],
    "CompuMethods": [
        ["boolean_CompuMethod", "/AUTOSAR_Platform/CompuMethods/boolean_CompuMethod", "IDENTICAL", "Platform boolean computation method"],
        ["CM_HornCmd_TextTable", "/HORN_CTRL/CompuMethods/CM_HornCmd_TextTable", "TEXTTABLE", "Horn command: OFF=0, ON=1"],
        ["CM_Volt_Linear", "/HORN_CTRL/CompuMethods/CM_Volt_Linear", "LINEAR", "PhysicalValue = Internal * 0.1V"],
        ["CM_Curr_Linear", "/HORN_CTRL/CompuMethods/CM_Curr_Linear", "LINEAR", "PhysicalValue = Internal * 0.01A"],
        ["CM_DutyRat_Linear", "/HORN_CTRL/CompuMethods/CM_DutyRat_Linear", "LINEAR", "PhysicalValue = Internal * 0.1%"],
        ["CM_HornTimes", "/HORN_CTRL/CompuMethods/CM_HornTimes", "IDENTICAL", "Horn repeat times identity"],
        ["CM_HornOperCmd_TextTable", "/HORN_CTRL/CompuMethods/CM_HornOperCmd_TextTable", "TEXTTABLE", "Horn operation command text table"],
    ],
    "CompuScales": [
        ["CM_HornCmd_TextTable", "0", "0", "OFF", "", "", "", ""],
        ["CM_HornCmd_TextTable", "1", "1", "ON", "", "", "", ""],
        ["CM_Volt_Linear", "", "", "", "1", "10", "0", ""],
        ["CM_Curr_Linear", "", "", "", "1", "100", "0", ""],
        ["CM_DutyRat_Linear", "", "", "", "1", "10", "0", ""],
        ["CM_HornOperCmd_TextTable", "0", "0", "IDLE", "", "", "", ""],
        ["CM_HornOperCmd_TextTable", "1", "1", "HONK", "", "", "", ""],
        ["CM_HornOperCmd_TextTable", "2", "2", "ALARM", "", "", "", ""],
    ],
    "DataConstrs": [
        ["boolean_DataConstr", "/AUTOSAR_Platform/DataConstrs/boolean_DataConstr", "0", "1", "Boolean constraint"],
        ["DC_App_HornCmd", "/HORN_CTRL/DataConstrs/DC_App_HornCmd", "0", "255", ""],
        ["DC_App_HornActSts", "/HORN_CTRL/DataConstrs/DC_App_HornActSts", "0", "255", ""],
        ["DC_App_Volt", "/HORN_CTRL/DataConstrs/DC_App_Volt", "0", "180", "0-18V in 0.1V steps"],
        ["DC_App_Curr", "/HORN_CTRL/DataConstrs/DC_App_Curr", "0", "500", "0-5A in 0.01A steps"],
        ["DC_App_DutyRat", "/HORN_CTRL/DataConstrs/DC_App_DutyRat", "0", "1000", "0-100% in 0.1% steps"],
    ],
    "SRInterfaces": [
        ["If_HornActSts_SR", "/HORN_CTRL/Interfaces/If_HornActSts_SR", "false", "Horn activation status signal"],
        ["If_Volt_SR", "/HORN_CTRL/Interfaces/If_Volt_SR", "false", "Vehicle voltage signal"],
        ["If_Curr_SR", "/HORN_CTRL/Interfaces/If_Curr_SR", "false", "Horn current signal"],
        ["If_FltSts_SR", "/HORN_CTRL/Interfaces/If_FltSts_SR", "false", "Fault status signal"],
        ["If_SrvOperSts_SR", "/HORN_CTRL/Interfaces/If_SrvOperSts_SR", "false", "Service operation status signal"],
    ],
    "SRDataElements": [
        ["If_HornActSts_SR", "ntfHornActSts", "/HORN_CTRL/ApplicationDataTypes/App_HornActSts", ""],
        ["If_Volt_SR", "ntfVolt", "/HORN_CTRL/ApplicationDataTypes/App_Volt", ""],
        ["If_Curr_SR", "ntfCurr", "/HORN_CTRL/ApplicationDataTypes/App_Curr", ""],
        ["If_FltSts_SR", "ntfFltSts", "/HORN_CTRL/ApplicationDataTypes/App_FltSts", ""],
        ["If_SrvOperSts_SR", "ntfSrvOperSts", "/HORN_CTRL/ApplicationDataTypes/App_SrvOperSts", ""],
    ],
    "CSInterfaces": [
        ["If_BOD_Horn_Atm_CS", "/HORN_CTRL/Interfaces/If_BOD_Horn_Atm_CS", "false", "Horn actuator CS interface"],
        ["If_BOD_HornCtrl_Enh_CS", "/HORN_CTRL/Interfaces/If_BOD_HornCtrl_Enh_CS", "false", "Horn control enhancement CS interface"],
        ["If_BOD_Horn_Dev_CS", "/HORN_CTRL/Interfaces/If_BOD_Horn_Dev_CS", "false", "Horn device CS interface"],
        ["getSrvOperSts", "/HORN_CTRL/Interfaces/getSrvOperSts", "false", "Get service operation status"],
    ],
    "CSOperations": [
        ["If_BOD_Horn_Atm_CS", "rrHornCmd", "Request horn command"],
        ["If_BOD_HornCtrl_Enh_CS", "rrHornCtrl", "Request horn control"],
        ["If_BOD_HornCtrl_Enh_CS", "getHornActSts", "Get horn activation status"],
        ["If_BOD_HornCtrl_Enh_CS", "getSrvOperSts", "Get service operation status"],
        ["If_BOD_Horn_Dev_CS", "rrHornOperCmd", "Request horn operation command"],
        ["getSrvOperSts", "getSrvOperSts", "Get service operation status"],
    ],
    "CSArguments": [
        ["If_BOD_Horn_Atm_CS", "rrHornCmd", "HornCmd", "IN", "/HORN_CTRL/ApplicationDataTypes/App_HornCmd", ""],
        ["If_BOD_Horn_Atm_CS", "rrHornCmd", "ReturnCode", "OUT", "/HORN_CTRL/ApplicationDataTypes/App_ReturnCode", ""],
        ["If_BOD_HornCtrl_Enh_CS", "rrHornCtrl", "HornCtrlMode", "IN", "/HORN_CTRL/ApplicationDataTypes/App_HornCtrlMode", ""],
        ["If_BOD_HornCtrl_Enh_CS", "rrHornCtrl", "ReturnCode", "OUT", "/HORN_CTRL/ApplicationDataTypes/App_ReturnCode", ""],
        ["If_BOD_HornCtrl_Enh_CS", "getHornActSts", "HornActSts", "OUT", "/HORN_CTRL/ApplicationDataTypes/App_HornActSts", ""],
        ["If_BOD_Horn_Dev_CS", "rrHornOperCmd", "HornOperCmd", "IN", "/HORN_CTRL/ApplicationDataTypes/App_HornOperCmd", ""],
        ["If_BOD_Horn_Dev_CS", "rrHornOperCmd", "DutyRat", "INOUT", "/HORN_CTRL/ApplicationDataTypes/App_DutyRat", ""],
        ["If_BOD_Horn_Dev_CS", "rrHornOperCmd", "ReturnCode", "OUT", "/HORN_CTRL/ApplicationDataTypes/App_ReturnCode", ""],
    ],
    "Ports": [
        # BOD_Horn_Atm (Atomic layer - horn actuator - Server)
        ["BOD_Horn_Atm", "Pp_BOD_Horn_Atm_CS", "P", "CS", "/HORN_CTRL/Interfaces/If_BOD_Horn_Atm_CS", "", "rrHornCmd", "SERVER-COM-SPEC", "", "1", "", "", "", "Server: handle horn command"],
        ["BOD_Horn_Atm", "Pp_HornActSts_SR", "P", "SR", "/HORN_CTRL/Interfaces/If_HornActSts_SR", "ntfHornActSts", "", "NONQUEUED-SENDER-COM-SPEC", "", "", "", "", "0", "Sender: horn status"],
        ["BOD_Horn_Atm", "Rp_Horn_Dev_CS", "R", "CS", "/HORN_CTRL/Interfaces/If_BOD_Horn_Dev_CS", "", "rrHornOperCmd", "CLIENT-COM-SPEC", "0", "", "", "", "", "Client: call horn device"],
        ["BOD_Horn_Atm", "Rp_Curr_SR", "R", "SR", "/HORN_CTRL/Interfaces/If_Curr_SR", "ntfCurr", "", "NONQUEUED-RECEIVER-COM-SPEC", "0", "", "false", "NONE", "0", "Receiver: current signal"],
        ["BOD_Horn_Atm", "Rp_FltSts_SR", "R", "SR", "/HORN_CTRL/Interfaces/If_FltSts_SR", "ntfFltSts", "", "NONQUEUED-RECEIVER-COM-SPEC", "0", "", "false", "NONE", "0", "Receiver: fault status"],
        ["BOD_Horn_Atm", "Rp_Volt_SR", "R", "SR", "/HORN_CTRL/Interfaces/If_Volt_SR", "ntfVolt", "", "NONQUEUED-RECEIVER-COM-SPEC", "0", "", "false", "NONE", "0", "Receiver: voltage signal"],
        # BOD_HornCtrl_Enh (Enhancement layer - Client/Server)
        ["BOD_HornCtrl_Enh", "Rp_Horn_Atm_CS", "R", "CS", "/HORN_CTRL/Interfaces/If_BOD_Horn_Atm_CS", "", "rrHornCmd", "CLIENT-COM-SPEC", "0", "", "", "", "", "Client: call horn actuator"],
        ["BOD_HornCtrl_Enh", "Rp_Horn_Atm_SR", "R", "SR", "/HORN_CTRL/Interfaces/If_HornActSts_SR", "ntfHornActSts", "", "NONQUEUED-RECEIVER-COM-SPEC", "0", "", "false", "NONE", "0", "Receiver: horn status"],
        ["BOD_HornCtrl_Enh", "Rp_Volt_SR", "R", "SR", "/HORN_CTRL/Interfaces/If_Volt_SR", "ntfVolt", "", "NONQUEUED-RECEIVER-COM-SPEC", "0", "", "false", "NONE", "120", "Receiver: voltage with init value"],
        ["BOD_HornCtrl_Enh", "Pp_BOD_HornCtrl_Enh_CS", "P", "CS", "/HORN_CTRL/Interfaces/If_BOD_HornCtrl_Enh_CS", "", "rrHornCtrl", "SERVER-COM-SPEC", "", "1", "", "", "", "Server: handle horn control"],
        ["BOD_HornCtrl_Enh", "Pp_SrvOperSts_SR", "P", "SR", "/HORN_CTRL/Interfaces/If_SrvOperSts_SR", "ntfSrvOperSts", "", "NONQUEUED-SENDER-COM-SPEC", "", "", "", "", "0", "Sender: service status"],
        # BOD_ChildLeftBehindAlert_Scen (Scenario layer)
        ["BOD_ChildLeftBehindAlert_Scen", "Rp_HornCtrl_Enh_CS", "R", "CS", "/HORN_CTRL/Interfaces/If_BOD_HornCtrl_Enh_CS", "", "rrHornCtrl", "CLIENT-COM-SPEC", "0", "", "", "", "", "Client: call horn control (multi-op port)"],
        ["BOD_ChildLeftBehindAlert_Scen", "Pp_Scen_getSrvOperSts", "P", "CS", "/HORN_CTRL/Interfaces/getSrvOperSts", "", "getSrvOperSts", "SERVER-COM-SPEC", "", "1", "", "", "", "Server: provide service op status"],
        ["BOD_ChildLeftBehindAlert_Scen", "Pp_Scen_SrvOperSts_SR", "P", "SR", "/HORN_CTRL/Interfaces/If_SrvOperSts_SR", "ntfSrvOperSts", "", "NONQUEUED-SENDER-COM-SPEC", "", "", "", "", "0", "Sender: service status"],
        ["BOD_ChildLeftBehindAlert_Scen", "Rp_HornActSts_SR", "R", "SR", "/HORN_CTRL/Interfaces/If_HornActSts_SR", "ntfHornActSts", "", "NONQUEUED-RECEIVER-COM-SPEC", "0", "", "false", "NONE", "0", "Receiver: horn status"],
        ["BOD_ChildLeftBehindAlert_Scen", "Rp_SrvOperSts_SR", "R", "SR", "/HORN_CTRL/Interfaces/If_SrvOperSts_SR", "ntfSrvOperSts", "", "NONQUEUED-RECEIVER-COM-SPEC", "0", "", "false", "NONE", "0", "Receiver: service status"],
        # BOD_SerchCar_Gen (Scenario layer - test component)
        ["BOD_SerchCar_Gen", "Rp_BOD_HornCtrl_Enh_CS", "R", "CS", "/HORN_CTRL/Interfaces/If_BOD_HornCtrl_Enh_CS", "", "rrHornCtrl", "CLIENT-COM-SPEC", "0", "", "", "", "", "Client: call horn control"],
        ["BOD_SerchCar_Gen", "Pp_Volt_SR", "P", "SR", "/HORN_CTRL/Interfaces/If_Volt_SR", "ntfVolt", "", "NONQUEUED-SENDER-COM-SPEC", "", "", "", "", "120", "Sender: voltage with init"],
    ],
    "Runnables": [
        ["BOD_Horn_Atm", "BOD_Horn_Atm_Init", "BOD_Horn_Atm_Init", "Init runnable"],
        ["BOD_Horn_Atm", "BOD_Horn_Atm_Step", "BOD_Horn_Atm_Step", "Periodic main step (10ms)"],
        ["BOD_Horn_Atm", "Runnable_CS", "Runnable_CS", "CS server: handle horn cmd"],
        ["BOD_HornCtrl_Enh", "BOD_HornCtrl_Enh_Init", "BOD_HornCtrl_Enh_Init", "Init runnable"],
        ["BOD_HornCtrl_Enh", "BOD_HornCtrl_Enh_Step", "BOD_HornCtrl_Enh_Step", "Periodic main step (10ms)"],
        ["BOD_HornCtrl_Enh", "Runnable_HornCtrl_CS", "Runnable_HornCtrl_CS", "CS server: handle horn control"],
        ["BOD_ChildLeftBehindAlert_Scen", "Scen_Init", "Scen_Init", "Init runnable"],
        ["BOD_ChildLeftBehindAlert_Scen", "Scen_Step", "Scen_Step", "Periodic main step (10ms)"],
        ["BOD_ChildLeftBehindAlert_Scen", "Runnable_NTF", "Runnable_NTF", "Notification handler for SR events"],
        ["BOD_SerchCar_Gen", "Gen_Init", "Gen_Init", "Init runnable"],
        ["BOD_SerchCar_Gen", "Gen_Step", "Gen_Step", "Periodic main step (20ms)"],
    ],
    "RunnableEvents": [
        # BOD_Horn_Atm
        ["BOD_Horn_Atm", "BOD_Horn_Atm_Init", "Init", "", "", "", "", ""],
        ["BOD_Horn_Atm", "BOD_Horn_Atm_Step", "Periodic", "10", "", "", "", ""],
        ["BOD_Horn_Atm", "Runnable_CS", "OperationInvoked", "", "Pp_BOD_Horn_Atm_CS", "rrHornCmd", "", ""],
        # BOD_HornCtrl_Enh
        ["BOD_HornCtrl_Enh", "BOD_HornCtrl_Enh_Init", "Init", "", "", "", "", ""],
        ["BOD_HornCtrl_Enh", "BOD_HornCtrl_Enh_Step", "Periodic", "10", "", "", "", ""],
        ["BOD_HornCtrl_Enh", "BOD_HornCtrl_Enh_Step", "DataReceived", "", "Rp_Horn_Atm_SR", "", "ntfHornActSts", ""],
        ["BOD_HornCtrl_Enh", "BOD_HornCtrl_Enh_Step", "DataReceived", "", "Rp_Volt_SR", "", "ntfVolt", ""],
        ["BOD_HornCtrl_Enh", "Runnable_HornCtrl_CS", "OperationInvoked", "", "Pp_BOD_HornCtrl_Enh_CS", "rrHornCtrl", "", ""],
        # BOD_ChildLeftBehindAlert_Scen
        ["BOD_ChildLeftBehindAlert_Scen", "Scen_Init", "Init", "", "", "", "", ""],
        ["BOD_ChildLeftBehindAlert_Scen", "Scen_Step", "Periodic", "10", "", "", "", ""],
        ["BOD_ChildLeftBehindAlert_Scen", "Runnable_NTF", "DataReceived", "", "Rp_HornActSts_SR", "", "ntfHornActSts", ""],
        ["BOD_ChildLeftBehindAlert_Scen", "Runnable_NTF", "DataReceived", "", "Rp_SrvOperSts_SR", "", "ntfSrvOperSts", ""],
        # BOD_SerchCar_Gen
        ["BOD_SerchCar_Gen", "Gen_Init", "Init", "", "", "", "", ""],
        ["BOD_SerchCar_Gen", "Gen_Step", "Periodic", "20", "", "", "", ""],
    ],
    "RunnableAccesses": [
        # ServerCallPoints (CS R-Port calls)
        ["BOD_HornCtrl_Enh", "BOD_HornCtrl_Enh_Step", "ServerCallPoint", "Rp_Horn_Atm_CS", "rrHornCmd", "", "SC_Rp_Horn_Atm_CS_rrHornCmd", "Call horn actuator"],
        ["BOD_ChildLeftBehindAlert_Scen", "Scen_Step", "ServerCallPoint", "Rp_HornCtrl_Enh_CS", "getHornActSts", "", "SC_Rp_HornCtrl_Enh_CS_getHornActSts", "Get horn status"],
        ["BOD_ChildLeftBehindAlert_Scen", "Scen_Step", "ServerCallPoint", "Rp_HornCtrl_Enh_CS", "getSrvOperSts", "", "SC_Rp_HornCtrl_Enh_CS_getSrvOperSts", "Get service status"],
        ["BOD_ChildLeftBehindAlert_Scen", "Scen_Step", "ServerCallPoint", "Rp_HornCtrl_Enh_CS", "rrHornCtrl", "", "SC_Rp_HornCtrl_Enh_CS_rrHornCtrl", "Call horn control"],
        ["BOD_SerchCar_Gen", "Gen_Step", "ServerCallPoint", "Rp_BOD_HornCtrl_Enh_CS", "rrHornCtrl", "", "SC_Rp_BOD_HornCtrl_Enh_CS_rrHornCtrl", "Call horn control"],
        # DataRead (SR R-Port read access)
        ["BOD_HornCtrl_Enh", "BOD_HornCtrl_Enh_Step", "DataRead", "Rp_Horn_Atm_SR", "", "ntfHornActSts", "DR_Rp_Horn_Atm_SR_ntfHornActSts", "Read horn status"],
        ["BOD_HornCtrl_Enh", "BOD_HornCtrl_Enh_Step", "DataRead", "Rp_Volt_SR", "", "ntfVolt", "DR_Rp_Volt_SR_ntfVolt", "Read voltage"],
        # DataWrite (SR P-Port write access)
        ["BOD_HornCtrl_Enh", "BOD_HornCtrl_Enh_Step", "DataWrite", "Pp_SrvOperSts_SR", "", "ntfSrvOperSts", "DW_Pp_SrvOperSts_SR_ntfSrvOperSts", "Write service status"],
    ],
    "CompositionConnectors": [
        ["Composition_HornCtrl", "Atm_Inst", "Pp_BOD_Horn_Atm_CS", "Enh_Inst", "Rp_Horn_Atm_CS", "Assembly", "CS: Atm → Enh"],
        ["Composition_HornCtrl", "Atm_Inst", "Pp_HornActSts_SR", "Enh_Inst", "Rp_Horn_Atm_SR", "Assembly", "SR: Atm horn status → Enh"],
        ["Composition_HornCtrl", "Atm_Inst", "Pp_HornActSts_SR", "Scen_Inst", "Rp_HornActSts_SR", "Assembly", "SR: Atm horn status → Scen"],
        ["Composition_HornCtrl", "Enh_Inst", "Pp_BOD_HornCtrl_Enh_CS", "Scen_Inst", "Rp_HornCtrl_Enh_CS", "Assembly", "CS: Enh → Scen"],
        ["Composition_HornCtrl", "Enh_Inst", "Pp_BOD_HornCtrl_Enh_CS", "Gen_Inst", "Rp_BOD_HornCtrl_Enh_CS", "Assembly", "CS: Enh → Gen"],
        ["Composition_HornCtrl", "Enh_Inst", "Pp_SrvOperSts_SR", "Scen_Inst", "Rp_SrvOperSts_SR", "Assembly", "SR: Enh service status → Scen"],
        ["Composition_HornCtrl", "Gen_Inst", "Pp_Volt_SR", "Enh_Inst", "Rp_Volt_SR", "Assembly", "SR: Gen voltage → Enh"],
    ],
    "Units": [
        ["Unit_Ampere", "/HORN_CTRL/Units/Unit_Ampere", "A", "1", "0", "Current unit (Ampere)"],
        ["Unit_Volt", "/HORN_CTRL/Units/Unit_Volt", "V", "1", "0", "Voltage unit (Volt)"],
        ["Unit_NoUnit", "/HORN_CTRL/Units/Unit_NoUnit", "-", "1", "0", "Dimensionless unit"],
    ],
}


def create_template_v2(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    for sheet_name, headers in SHEETS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for row in EXAMPLE_ROWS.get(sheet_name, []):
            sheet.append(row)
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column) + 2
            sheet.column_dimensions[column[0].column_letter].width = min(max(width, 14), 64)
    _add_validations(workbook)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _add_validations(workbook: Workbook) -> None:
    rules = {
        "Components": {"B": '"Application,Composition"'},
        "PrimitiveDataTypes": {"E": '"boolean,uint8,uint16,uint32,uint64,sint8,sint16,sint32,float32"', "H": '"READ-ONLY,READ-WRITE,NOT-ACCESSIBLE"'},
        "CompuMethods": {"C": '"TEXTTABLE,LINEAR,IDENTICAL"'},
        "SRInterfaces": {"C": '"true,false"'},
        "CSInterfaces": {"C": '"true,false"'},
        "CSArguments": {"D": '"IN,OUT,INOUT"'},
        "Ports": {"C": '"P,R"', "D": '"SR,CS"', "H": '"CLIENT-COM-SPEC,SERVER-COM-SPEC,NONQUEUED-SENDER-COM-SPEC,NONQUEUED-RECEIVER-COM-SPEC,QUEUED-SENDER-COM-SPEC,QUEUED-RECEIVER-COM-SPEC"'},
        "RunnableEvents": {"C": '"Init,Periodic,OperationInvoked,DataReceived"'},
        "RunnableAccesses": {"C": '"DataRead,DataWrite,ServerCallPoint"'},
        "CompositionConnectors": {"F": '"Assembly"'},
    }
    for sheet_name, column_rules in rules.items():
        sheet = workbook[sheet_name]
        for column, formula in column_rules.items():
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            sheet.add_data_validation(validation)
            validation.add(f"{column}2:{column}1000")
