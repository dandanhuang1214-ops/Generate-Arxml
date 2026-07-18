import arxml_codegen.cli as cli_module
from arxml_codegen.cli import build_parser
from arxml_codegen.excel.template import create_template_v2
from arxml_codegen.excel.reader import load_workbook_v2
from arxml_codegen.generator.arxml_writer import (
    GeneratorConfig,
    build_arxml_v2,
    validate_model_v2,
    write_outputs,
)
from arxml_codegen.models.schema import (
    CompuMethodRow,
    CompuScaleRow,
    ComponentPrototypeRow,
    ComponentV2Row,
    CompositionConnectorV2Row,
    CSArgumentRow,
    CSInterfaceRow,
    CSOperationRow,
    DataConstrRow,
    DataTypeMappingRow,
    PortV2Row,
    PortRecordInitValueRow,
    PrimitiveDataTypeRow,
    RecordElementRow,
    RecordTypeRow,
    RunnableAccessRow,
    RunnableV2Row,
    SRDataElementRow,
    SRInterfaceRow,
    UnitRow,
    WorkbookV2Model,
)
from arxml_codegen.validator.engine import run_all as run_core_validation


def _base_model() -> WorkbookV2Model:
    return WorkbookV2Model(
        components=[
            ComponentV2Row("Components", 2, "Enh", "Application", "/ComponentTypes", "", ""),
            ComponentV2Row("Components", 3, "Atm", "Application", "/ComponentTypes", "", ""),
            ComponentV2Row("Components", 4, "Composition_Test", "Composition", "/System", "", ""),
        ],
        primitive_data_types=[
            PrimitiveDataTypeRow("PrimitiveDataTypes", 2, "App_Bool", "/DataTypes/App_Bool", "uint8", "/Platform/uint8", "boolean", "", "", "READ-ONLY"),
        ],
        sr_interfaces=[
            SRInterfaceRow("SRInterfaces", 2, "If_Bool_SR", "/Interfaces/If_Bool_SR", "false"),
        ],
        sr_data_elements=[
            SRDataElementRow("SRDataElements", 2, "If_Bool_SR", "ntfBool", "/DataTypes/App_Bool"),
        ],
        cs_interfaces=[
            CSInterfaceRow("CSInterfaces", 2, "If_Cmd_CS", "/Interfaces/If_Cmd_CS", "false"),
        ],
        cs_operations=[
            CSOperationRow("CSOperations", 2, "If_Cmd_CS", "rrCmd"),
        ],
        cs_arguments=[
            CSArgumentRow("CSArguments", 2, "If_Cmd_CS", "rrCmd", "CmdVal", "IN", "/DataTypes/App_Bool"),
        ],
        ports=[
            PortV2Row("Ports", 2, "Atm", "pBool", "P", "SR", "/Interfaces/If_Bool_SR", "ntfBool", "", "", "", "", "", "", ""),
            PortV2Row("Ports", 3, "Enh", "rBool", "R", "SR", "/Interfaces/If_Bool_SR", "ntfBool", "", "", "", "", "", "", ""),
        ],
        runnables=[
            RunnableV2Row("Runnables", 2, "Enh", "Enh_Init", "Enh_Init"),
        ],
        component_prototypes=[
            ComponentPrototypeRow("ComponentPrototypes", 2, "Composition_Test", "Atm_Inst", "Atm", "/ComponentTypes/Atm"),
            ComponentPrototypeRow("ComponentPrototypes", 3, "Composition_Test", "Enh_Inst", "Enh", "/ComponentTypes/Enh"),
        ],
        composition_connectors=[
            CompositionConnectorV2Row("CompositionConnectors", 2, "Composition_Test", "Atm_Inst", "pBool", "Enh_Inst", "rBool", "Assembly"),
        ],
    )


def test_parser_defaults() -> None:
    parser = build_parser()
    args = parser.parse_args([])
    assert args.config.parts[-2:] == ("config", "project.yaml")
    assert args.dry_run is False
    assert args.create_template is None


def test_cli_blocks_generation_when_core_validation_has_errors(monkeypatch, tmp_path) -> None:
    model = _base_model()
    model.ports[0].init_value = "not-a-number"
    model.ports[0].init_value_type = "Numeric"
    config = GeneratorConfig(
        workbook=tmp_path / "input.xlsx",
        output=tmp_path / "output.arxml",
        report=tmp_path / "report.md",
        matlab_init=None,
    )
    generated = False

    def mark_generated(*args, **kwargs) -> None:
        nonlocal generated
        generated = True

    monkeypatch.setattr(cli_module, "load_config", lambda path: config)
    monkeypatch.setattr(cli_module, "load_workbook_v2", lambda path: model)
    monkeypatch.setattr(cli_module, "write_outputs", mark_generated)
    monkeypatch.setattr("sys.argv", ["arxml-codegen"])

    assert cli_module.main() == 1
    assert generated is False


def test_generation_report_contains_core_findings(tmp_path) -> None:
    model = _base_model()
    findings = run_core_validation(model)
    config = GeneratorConfig(
        workbook=tmp_path / "input.xlsx",
        output=tmp_path / "output.arxml",
        report=tmp_path / "report.md",
        matlab_init=None,
    )

    write_outputs(model, config, [], findings)
    report = config.report.read_text(encoding="utf-8")

    assert "CORE-" in report
    assert config.output.with_suffix(".arxml.manifest.json").exists()


def test_linear_compu_method_writes_limits_and_unit_ref() -> None:
    model = _base_model()
    model.primitive_data_types[0].application_type_name = "App_DutyRat"
    model.primitive_data_types[0].application_type_path = "/DataTypes/App_DutyRat"
    model.primitive_data_types[0].compu_method_ref = "/DataTypes/CompuMethods/CM_App_DutyRat_Linear"
    model.primitive_data_types[0].unit_ref = "/DataTypes/Units/No_Unit"
    model.compu_methods = [
        CompuMethodRow(
            "CompuMethods",
            2,
            "CM_App_DutyRat_Linear",
            "/DataTypes/CompuMethods/CM_App_DutyRat_Linear",
            "LINEAR",
        )
    ]
    model.compu_scales = [
        CompuScaleRow(
            "CompuScales",
            2,
            "CM_App_DutyRat_Linear",
            "0",
            "100",
            "",
            "0.1",
            "1",
            "0",
        )
    ]

    xml_text = etree_to_text(build_arxml_v2(model))

    assert "<UNIT-REF DEST=\"UNIT\">/DataTypes/Units/No_Unit</UNIT-REF>" in xml_text
    assert "<LOWER-LIMIT INTERVAL-TYPE=\"CLOSED\">0</LOWER-LIMIT>" in xml_text
    assert "<UPPER-LIMIT INTERVAL-TYPE=\"CLOSED\">100</UPPER-LIMIT>" in xml_text


def test_runnable_accesses_match_davinci_child_order() -> None:
    model = _base_model()
    model.ports.append(
        PortV2Row(
            "Ports",
            4,
            "Enh",
            "pBool",
            "P",
            "SR",
            "/Interfaces/If_Bool_SR",
            "ntfBool",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        )
    )
    model.ports.append(
        PortV2Row(
            "Ports",
            5,
            "Enh",
            "rCmd",
            "R",
            "CS",
            "/Interfaces/If_Cmd_CS",
            "",
            "rrCmd",
            "CLIENT-COM-SPEC",
            "",
            "",
            "",
            "",
            "",
        )
    )
    model.runnables[0].runnable_name = "Enh_Step"
    model.runnables[0].symbol = "Enh_Step"
    model.runnable_accesses = [
        RunnableAccessRow("RunnableAccesses", 2, "Enh", "Enh_Step", "DataRead", "rBool", "ntfBool", "", ""),
        RunnableAccessRow("RunnableAccesses", 3, "Enh", "Enh_Step", "DataWrite", "pBool", "ntfBool", "", ""),
        RunnableAccessRow("RunnableAccesses", 4, "Enh", "Enh_Step", "ServerCallPoint", "rCmd", "", "rrCmd", ""),
    ]

    xml_text = etree_to_text(build_arxml_v2(model))
    runnable_text = xml_text[
        xml_text.index("<SHORT-NAME>Enh_Step</SHORT-NAME>") : xml_text.index("</RUNNABLE-ENTITY>", xml_text.index("<SHORT-NAME>Enh_Step</SHORT-NAME>"))
    ]

    assert runnable_text.index("<DATA-RECEIVE-POINT-BY-ARGUMENTS>") < runnable_text.index("<DATA-SEND-POINTS>")
    assert runnable_text.index("<DATA-SEND-POINTS>") < runnable_text.index("<SERVER-CALL-POINTS>")
    assert runnable_text.index("<SERVER-CALL-POINTS>") < runnable_text.index("<SYMBOL>Enh_Step</SYMBOL>")


def test_template_validations_are_present(tmp_path) -> None:
    path = tmp_path / "template.xlsx"
    create_template_v2(path)
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    assert workbook["Components"].data_validations.count > 0
    assert workbook["Ports"].data_validations.count > 0
    assert workbook["RunnableEvents"].data_validations.count > 0


def test_empty_component_name_reported() -> None:
    model = _base_model()
    model.components[0].component_name = ""
    errors = validate_model_v2(model)
    assert any("ComponentName" in error for error in errors)


def test_invalid_base_type_reported() -> None:
    model = _base_model()
    model.primitive_data_types[0].base_type = "unknown_type"
    errors = validate_model_v2(model)
    assert any("BaseType" in error for error in errors)


def test_unknown_component_kind_reported() -> None:
    model = _base_model()
    model.components[0].component_kind = "InvalidKind"
    errors = validate_model_v2(model)
    assert any("ComponentKind" in error for error in errors)


def test_v2_template_loads_and_generates_arxml(tmp_path) -> None:
    path = tmp_path / "hornctrl_v2.xlsx"
    create_template_v2(path)
    model = load_workbook_v2(path)
    errors = validate_model_v2(model)
    assert errors == []

    tree = build_arxml_v2(model)
    text = str(tree.getroot().xpath("count(//*[local-name()='SW-COMPONENT-PROTOTYPE'])"))
    assert text == "4.0"


def test_v2_connector_uses_prototype_context(tmp_path) -> None:
    path = tmp_path / "hornctrl_v2.xlsx"
    create_template_v2(path)
    model = load_workbook_v2(path)
    tree = build_arxml_v2(model)
    xml_text = etree_to_text(tree)
    assert "/HORN_CTRL/System/Composition_HornCtrl/Atm_Inst" in xml_text
    assert "/HORN_CTRL/System/Composition_HornCtrl/Enh_Inst" in xml_text
    assert "/HORN_CTRL/System/Composition_HornCtrl/Gen_Inst" in xml_text


def test_v2_writes_delegation_connector_and_outer_composition_port() -> None:
    model = _base_model()
    model.composition_connectors.append(
        CompositionConnectorV2Row(
            "CompositionConnectors",
            3,
            "Composition_Test",
            "Composition_Test",
            "pBool_Out",
            "Atm_Inst",
            "pBool",
            "Delegation",
        )
    )

    assert validate_model_v2(model) == []

    tree = build_arxml_v2(model)
    xml_text = etree_to_text(tree)

    assert "DELEGATION-SW-CONNECTOR" in xml_text
    assert "<SHORT-NAME>pBool_Out</SHORT-NAME>" in xml_text
    assert "OUTER-PORT-REF" in xml_text
    assert "INNER-PORT-IREF" in xml_text


def test_v2_template_covers_record_and_linear_compu(tmp_path) -> None:
    path = tmp_path / "hornctrl_v2.xlsx"
    create_template_v2(path)
    model = load_workbook_v2(path)
    tree = build_arxml_v2(model)
    xml_text = etree_to_text(tree)
    assert "APPLICATION-RECORD-DATA-TYPE" in xml_text
    assert "IMPLEMENTATION-DATA-TYPE" in xml_text
    assert "CM_Volt_Linear" in xml_text
    assert "COMPU-RATIONAL-COEFFS" in xml_text


def test_v2_reports_unknown_data_type_ref(tmp_path) -> None:
    path = tmp_path / "hornctrl_v2.xlsx"
    create_template_v2(path)
    model = load_workbook_v2(path)
    model.sr_data_elements[0].application_type_ref = "/HORN_CTRL/ApplicationDataTypes/MissingType"
    errors = validate_model_v2(model)
    assert any("SRDataElements!R2 ApplicationTypeRef" in error for error in errors)


def test_v2_includes_units(tmp_path) -> None:
    path = tmp_path / "hornctrl_v2.xlsx"
    create_template_v2(path)
    model = load_workbook_v2(path)
    tree = build_arxml_v2(model)
    xml_text = etree_to_text(tree)
    assert "UNIT" in xml_text
    assert "FACTOR-SI-TO-UNIT" in xml_text


def test_init_value_type_boolean_writes_text_value_spec() -> None:
    model = _base_model()
    model.ports[0].init_value = "1"
    model.ports[0].init_value_type = "Boolean"

    tree = build_arxml_v2(model)
    xml_text = etree_to_text(tree)

    assert "<CATEGORY>BOOLEAN</CATEGORY>" in xml_text
    assert "<V>1</V>" in xml_text


def test_core_reports_enum_init_value_numeric_literal() -> None:
    model = _base_model()
    model.primitive_data_types[0].application_type_name = "App_Mode"
    model.primitive_data_types[0].application_type_path = "/DataTypes/App_Mode"
    model.primitive_data_types[0].compu_method_ref = "/CompuMethods/CM_Mode"
    model.sr_data_elements[0].application_type_ref = "/DataTypes/App_Mode"
    model.compu_methods = [
        CompuMethodRow("CompuMethods", 2, "CM_Mode", "/CompuMethods/CM_Mode", "TEXTTABLE"),
    ]
    model.compu_scales = [
        CompuScaleRow("CompuScales", 2, "CM_Mode", "0", "0", "OFF", "", "", ""),
        CompuScaleRow("CompuScales", 3, "CM_Mode", "1", "1", "ON", "", "", ""),
    ]
    model.ports[0].init_value = "1"
    model.ports[0].init_value_type = "Enum"

    findings = run_core_validation(model)

    assert any(finding.code == "CORE-010-INIT-VALUE-TYPE-MISMATCH" for finding in findings)


def test_core_reports_application_type_declared_as_primitive_and_record() -> None:
    model = _base_model()
    model.record_types = [
        RecordTypeRow(
            "RecordTypes",
            2,
            "App_Bool",
            "/DataTypes/App_Bool",
            "Impl_Bool",
            "/DataTypes/Impl_Bool",
        ),
    ]

    codes = {finding.code for finding in run_core_validation(model)}

    assert "CORE-010-DATATYPE-KIND-CONFLICT" in codes


def test_record_init_value_writes_record_value_specification() -> None:
    model = _base_model()
    model.record_types = [
        RecordTypeRow("RecordTypes", 2, "App_Record", "/DataTypes/App_Record", "Impl_Record", "/Impl/Impl_Record"),
    ]
    model.record_elements = [
        RecordElementRow("RecordElements", 2, "App_Record", "FieldA", "/DataTypes/App_Bool", "/Platform/uint8", "1"),
        RecordElementRow("RecordElements", 3, "App_Record", "FieldB", "/DataTypes/App_Bool", "/Platform/uint8", "2"),
    ]
    model.sr_data_elements[0].application_type_ref = "/DataTypes/App_Record"
    model.ports[0].init_value_type = "Record"
    model.port_record_init_values = [
        PortRecordInitValueRow("PortRecordInitValues", 2, "Atm", "pBool", "FieldA", "1", "Boolean"),
        PortRecordInitValueRow("PortRecordInitValues", 3, "Atm", "pBool", "FieldB", "0", "Boolean"),
    ]

    tree = build_arxml_v2(model)
    xml_text = etree_to_text(tree)

    assert "RECORD-VALUE-SPECIFICATION" in xml_text
    assert "FieldA" in xml_text
    assert "FieldB" in xml_text


def test_nested_record_init_value_writes_recursive_record_specification() -> None:
    model = _base_model()
    model.record_types = [
        RecordTypeRow("RecordTypes", 2, "App_Outer", "/DataTypes/App_Outer", "Impl_Outer", "/DataTypes/Impl_Outer"),
        RecordTypeRow("RecordTypes", 3, "App_Inner", "/DataTypes/App_Inner", "Impl_Inner", "/DataTypes/Impl_Inner"),
    ]
    model.record_elements = [
        RecordElementRow(
            "RecordElements",
            2,
            "App_Outer",
            "Nested",
            "/DataTypes/App_Inner",
            "/DataTypes/Impl_Inner",
            "1",
            "NestedImpl",
            "Record",
        ),
        RecordElementRow(
            "RecordElements",
            3,
            "App_Inner",
            "Value",
            "/DataTypes/App_Bool",
            "/Platform/uint8",
            "1",
            "RawValue",
            "Value",
        ),
    ]
    model.sr_data_elements[0].application_type_ref = "/DataTypes/App_Outer"
    model.ports[0].init_value_type = "Record"
    model.port_record_init_values = [
        PortRecordInitValueRow(
            "PortRecordInitValues",
            2,
            "Atm",
            "pBool",
            "Nested.Value",
            "1",
            "Boolean",
        ),
    ]

    xml_text = etree_to_text(build_arxml_v2(model))
    findings = run_core_validation(model)

    assert xml_text.count("<RECORD-VALUE-SPECIFICATION>") == 2
    assert "<SHORT-LABEL>Nested</SHORT-LABEL>" in xml_text
    assert "<SHORT-LABEL>Value</SHORT-LABEL>" in xml_text
    assert not any(
        finding.code == "CORE-010-INIT-VALUE-RECORD-INCOMPLETE"
        for finding in findings
    )


def test_core_reports_incomplete_record_init_value() -> None:
    model = _base_model()
    model.record_types = [
        RecordTypeRow("RecordTypes", 2, "App_Record", "/DataTypes/App_Record", "Impl_Record", "/Impl/Impl_Record"),
    ]
    model.record_elements = [
        RecordElementRow("RecordElements", 2, "App_Record", "FieldA", "/DataTypes/App_Bool", "/Platform/uint8", "1"),
        RecordElementRow("RecordElements", 3, "App_Record", "FieldB", "/DataTypes/App_Bool", "/Platform/uint8", "2"),
    ]
    model.sr_data_elements[0].application_type_ref = "/DataTypes/App_Record"
    model.ports[0].init_value_type = "Record"
    model.port_record_init_values = [
        PortRecordInitValueRow("PortRecordInitValues", 2, "Atm", "pBool", "FieldA", "1", "Boolean"),
    ]

    findings = run_core_validation(model)

    assert any(finding.code == "CORE-010-INIT-VALUE-RECORD-INCOMPLETE" for finding in findings)


def test_core_reports_compu_scale_gap_and_overlap() -> None:
    model = WorkbookV2Model(
        compu_scales=[
            CompuScaleRow("CompuScales", 2, "CM_Mode", "0", "1", "OFF", "", "", ""),
            CompuScaleRow("CompuScales", 3, "CM_Mode", "3", "4", "ON", "", "", ""),
            CompuScaleRow("CompuScales", 4, "CM_Overlap", "0", "2", "A", "", "", ""),
            CompuScaleRow("CompuScales", 5, "CM_Overlap", "2", "3", "B", "", "", ""),
        ],
    )

    codes = {finding.code for finding in run_core_validation(model)}

    assert "CORE-010-COMPU-SCALE-GAP" in codes
    assert "CORE-010-COMPU-SCALE-OVERLAP" in codes


def test_core_reports_unknown_local_compu_method_reference() -> None:
    model = WorkbookV2Model(
        primitive_data_types=[
            PrimitiveDataTypeRow(
                "PrimitiveDataTypes",
                2,
                "App_uint16",
                "/DataTypes/App_uint16",
                "uint16",
                "/AUTOSAR_Platform/ImplementationDataTypes/uint16",
                "uint16",
                "/DataTypes/CompuMethods/CM_App_uint16_Identical",
                "/AUTOSAR_Platform/DataConstrs/uint16_DataConstr",
                "READ-ONLY",
            )
        ]
    )

    codes = {finding.code for finding in run_core_validation(model)}

    assert "CORE-010-COMPU-METHOD-REF-UNKNOWN" in codes


def test_core_reports_linear_physical_range_inconsistency() -> None:
    model = WorkbookV2Model(
        primitive_data_types=[
            PrimitiveDataTypeRow(
                "PrimitiveDataTypes",
                2,
                "App_DutyRat",
                "/DataTypes/App_DutyRat",
                "uint8",
                "/AUTOSAR_Platform/ImplementationDataTypes/uint8",
                "uint8",
                "/DataTypes/CompuMethods/CM_App_DutyRat_Linear",
                "/DataTypes/DataConstrs/DC_App_DutyRat",
                "READ-ONLY",
            )
        ],
        compu_methods=[
            CompuMethodRow(
                "CompuMethods",
                2,
                "CM_App_DutyRat_Linear",
                "/DataTypes/CompuMethods/CM_App_DutyRat_Linear",
                "LINEAR",
            )
        ],
        compu_scales=[
            CompuScaleRow(
                "CompuScales",
                2,
                "CM_App_DutyRat_Linear",
                "0",
                "100",
                "",
                "0.1",
                "1",
                "0",
            )
        ],
        data_constrs=[
            DataConstrRow(
                "DataConstrs",
                2,
                "DC_App_DutyRat",
                "/DataTypes/DataConstrs/DC_App_DutyRat",
                "0",
                "255",
            )
        ],
    )

    codes = {finding.code for finding in run_core_validation(model)}

    assert "CORE-010-PHYS-RANGE-CONSISTENCY" in codes


def test_core_reports_dataconstr_not_covered_by_compu_scales() -> None:
    model = WorkbookV2Model(
        primitive_data_types=[
            PrimitiveDataTypeRow(
                "PrimitiveDataTypes",
                2,
                "App_Mode",
                "/DataTypes/App_Mode",
                "uint8",
                "/Platform/uint8",
                "uint8",
                "/CompuMethods/CM_Mode",
                "/DataConstrs/DC_Mode",
                "READ-ONLY",
            ),
        ],
        compu_methods=[
            CompuMethodRow("CompuMethods", 2, "CM_Mode", "/CompuMethods/CM_Mode", "TEXTTABLE"),
        ],
        compu_scales=[
            CompuScaleRow("CompuScales", 2, "CM_Mode", "0", "1", "OFF", "", "", ""),
        ],
        data_constrs=[
            DataConstrRow("DataConstrs", 2, "DC_Mode", "/DataConstrs/DC_Mode", "0", "3"),
        ],
    )

    codes = {finding.code for finding in run_core_validation(model)}

    assert "CORE-010-DATACONSTR-COVERAGE" in codes


def test_core_reports_unknown_unit_and_missing_mapping() -> None:
    model = WorkbookV2Model(
        primitive_data_types=[
            PrimitiveDataTypeRow(
                "PrimitiveDataTypes",
                2,
                "App_Speed",
                "/DataTypes/App_Speed",
                "uint16",
                "/Platform/uint16",
                "uint16",
                "",
                "",
                "READ-ONLY",
                "/Units/Unit_kmh",
            ),
        ],
        units=[
            UnitRow("Units", 2, "Unit_V", "/Units/Unit_V", "V", "1", "0"),
        ],
        sr_data_elements=[
            SRDataElementRow("SRDataElements", 2, "If_Speed", "Speed", "/DataTypes/App_Speed"),
        ],
        cs_arguments=[
            CSArgumentRow("CSArguments", 2, "If_Cmd", "rrCmd", "Arg", "IN", "/DataTypes/App_Arg"),
        ],
        record_elements=[
            RecordElementRow("RecordElements", 2, "App_Record", "Field", "/DataTypes/App_Field", "/Platform/uint8", "1"),
        ],
        data_type_mappings=[
            DataTypeMappingRow("DataTypeMappings", 2, "/Mapping", "/DataTypes/Other", "/Platform/uint8"),
        ],
    )

    codes = {finding.code for finding in run_core_validation(model)}

    assert "CORE-010-UNIT-UNKNOWN" in codes
    assert "CORE-010-MAPPING-MISSING" in codes


def etree_to_text(tree) -> str:
    from lxml import etree

    return etree.tostring(tree, encoding="unicode")
