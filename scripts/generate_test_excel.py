"""Generate arxml_input_test.xlsx from the wiper/washer system definition in test.md"""
from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


SHEETS: dict[str, list[str]] = {
    "Cover": [],
    "Components": ["ComponentName", "ComponentKind", "PackagePath", "IsComposition", "Description"],
    "DataTypes": ["ADTName", "IDTName", "BaseType", "IsEnum", "CompuMethod", "ValueDefinition", "Description"],
    "PortInterfaces": ["InterfaceName", "InterfaceKind", "DataElementName", "DataTypeADT", "OperationName", "Description"],
    "Operations": ["InterfaceName", "OperationName", "ArgumentName", "ArgumentDirection", "ArgumentADT", "Description"],
    "Ports": ["ComponentName", "PortName", "PortDirection", "InterfaceKind", "InterfaceName", "DataElementName", "OperationName", "InitValue", "ComSpecType", "Description"],
    "Runnables": ["ComponentName", "RunnableName", "Symbol", "Description"],
    "RunnableEvents": ["ComponentName", "RunnableName", "TriggerType", "PeriodMs", "PortName", "OperationName", "DataElementName", "AccessType", "Description"],
    "CompositionConnectors": ["CompositionName", "ProviderComponent", "ProviderPort", "RequesterComponent", "RequesterPort", "ConnectorType", "Description"],
}

OUTPUT = Path("D:/work/SOA/code/data/input/WW_SWC_Design.xlsx")


# ============================================================================
# Cover — meta sheet describing the workbook
# ============================================================================
COVER_META = {
    "project": "WW — 雨刮/洗涤控制系统",
    "version": "0.1.0",
    "date": "2026-05-20",
    "author": "SOA Team",
    "target": "AUTOSAR Classic Platform 4.3.0",
    "tool": "arxml-codegen",
    "description": (
        "本工作簿定义了雨刮/洗涤控制系统的完整 AUTOSAR SWC 架构，包含：\n"
        "  - 1 个 Composition（WW_Comp）+ 1 个增强层 SWC（WW_Enh）+ 11 个原子层 SWC（BOD_*_Atm）\n"
        "  - 36 路 S/R 输入信号（HWA 硬件 / CAN 总线 / EPRM 存储 / 整车状态）\n"
        "  - 22 路 S/R 输出信号（硬件驱动 + 状态反馈）\n"
        "  - 11 组 C/S 服务接口（执行类 6 组 + 状态上报类 5 组）\n"
        "工作流：填写本表 → run_codegen.ps1 → 生成 ARXML → DaVinci Developer 导入验证。"
    ),
    "naming": (
        "命名规则：\n"
        "  接口：rr* = C/S 服务接口，if* = S/R 信号接口\n"
        "  端口：r* = R-Port（接收端），p* = P-Port（提供端）\n"
        "  组件：WW_Enh = 增强/控制层，BOD_*_Atm = 原子/执行层，WW_Comp = 顶层组合\n"
        "  信号类型：Vb* = boolean，Ve* = enum/uint，Vu* = config/calibration"
    ),
}

SHEET_DESCRIPTIONS = [
    ("Cover", "封面与本表 — 工作簿概述、架构说明、Sheet 索引"),
    ("Components", "组件定义 — 13 个 SWC（1 Composition + 1 Enh + 11 Atm），含包路径和组件类型"),
    ("DataTypes", "数据类型 — 32 个 ADT/IDT 定义，含枚举文本表和 CompuMethod 映射"),
    ("PortInterfaces", "端口接口 — 69 个（11 C/S + 36 输入 S/R + 22 输出 S/R）"),
    ("Operations", "操作参数 — 33 行 C/S 操作入参/出参定义"),
    ("Ports", "端口实例 — 102 个（C/S 22 + 输入 S/R 36 + 输出 S/R 44）"),
    ("Runnables", "运行实体 — 24 个（Enh 周期任务 + 各 Atm 初始化与服务处理）"),
    ("RunnableEvents", "触发事件与端口访问 — TriggerType（Init/Periodic/OperationInvoked/DataReceived）+ AccessType（Read/Write/Call）"),
    ("CompositionConnectors", "组合连接器 — 33 个（C/S 11 组 + S/R 反馈 22 组）"),
]


def build_cover():
    """Return cover content as list of (row_type, data) tuples for special rendering."""
    rows = []

    # Title block
    rows.append(("title", "WW 雨刮/洗涤控制系统 — SWC 设计工作簿"))
    rows.append(("blank", ""))
    rows.append(("meta", "项目名称", COVER_META["project"]))
    rows.append(("meta", "版本", COVER_META["version"]))
    rows.append(("meta", "日期", COVER_META["date"]))
    rows.append(("meta", "作者", COVER_META["author"]))
    rows.append(("meta", "目标平台", COVER_META["target"]))
    rows.append(("meta", "生成工具", COVER_META["tool"]))
    rows.append(("blank", ""))

    # Architecture overview
    rows.append(("section", "系统架构概述"))
    rows.append(("text", COVER_META["description"]))
    rows.append(("blank", ""))

    # Sheet index
    rows.append(("section", "Sheet 索引"))
    rows.append(("index_header", ""))  # special marker for the column header row
    for sheet_name, desc in SHEET_DESCRIPTIONS[1:]:  # skip Cover itself in index
        rows.append(("index_row", sheet_name, desc))
    rows.append(("blank", ""))

    # Naming conventions
    rows.append(("section", "命名约定"))
    rows.append(("text", COVER_META["naming"]))

    return rows

def build_components():
    return [
        ["WW_Comp", "Composition", "/ComponentTypes", "TRUE", "Top composition – wiper/washer system"],
        ["WW_Enh", "Application", "/ComponentTypes", "FALSE", "Enhancement layer – wiper/washer control logic"],
        ["BOD_FWiper_Atm", "Application", "/ComponentTypes", "FALSE", "Atomic layer – front wiper motor drive"],
        ["BOD_FWasher_Atm", "Application", "/ComponentTypes", "FALSE", "Atomic layer – front washer motor drive"],
        ["BOD_RWiper_Atm", "Application", "/ComponentTypes", "FALSE", "Atomic layer – rear wiper motor drive"],
        ["BOD_RWasher_Atm", "Application", "/ComponentTypes", "FALSE", "Atomic layer – rear washer motor drive"],
        ["BOD_FWindHeater_Atm", "Application", "/ComponentTypes", "FALSE", "Atomic layer – front windshield heater"],
        ["BOD_FNozzleHeater_Atm", "Application", "/ComponentTypes", "FALSE", "Atomic layer – front nozzle heater"],
        ["BOD_WiperStatus_Atm", "Application", "/ComponentTypes", "FALSE", "Atomic layer – wiper status reporting"],
        ["BOD_RainSensor_Atm", "Application", "/ComponentTypes", "FALSE", "Atomic layer – rain sensor status"],
        ["BOD_Maintenance_Atm", "Application", "/ComponentTypes", "FALSE", "Atomic layer – maintenance status"],
        ["BOD_WashLiquid_Atm", "Application", "/ComponentTypes", "FALSE", "Atomic layer – wash liquid status"],
        ["BOD_HeaterStatus_Atm", "Application", "/ComponentTypes", "FALSE", "Atomic layer – heater status"],
    ]


# ============================================================================
# DataTypes
# ============================================================================
def build_data_types():
    rows = []

    # ---- CS 内部用的枚举类型（有 CompuMethod + TEXTTABLE） ----
    enums = [
        ("ADT_ReturnCode", "IDT_ReturnCode", "uint8", "ReturnCode_CompuMethod",
         "0:SUCCESS;1:FAILURE;2:FAIL_UNAVAILABLE;3:FAIL_INVALID_PARAM", "Common operation return code"),
        ("ADT_FWiperCmd", "IDT_FWiperCmd", "uint8", "FWiperCmd_CompuMethod",
         "0:Stop;1:Low;2:High", "Front wiper speed command"),
        ("ADT_FWiperSts", "IDT_FWiperSts", "uint8", "FWiperSts_CompuMethod",
         "0:OFF;1:LS;2:HS;3:Wash;4:Maint;5:AUTO;6:INT;7:MIST", "Front wiper run status"),
        ("ADT_FWiperSWSts", "IDT_FWiperSWSts", "uint8", "FWiperSWSts_CompuMethod",
         "0:OFF;1:MIST;2:LS;3:HS;4:AUTO;5:INT;6:Reserved;7:Invalid", "Front wiper switch status"),
        ("ADT_RainSensitivity", "IDT_RainSensitivity", "uint8", "RainSensitivity_CompuMethod",
         "0:Most Insensitive;1:More Insensitive;2:Insensitive;3:Normal;4:Sensitive;5:More Sensitive;6:Most Sensitive;7:Reserved",
         "Rain sensor sensitivity level"),
    ]
    for adt, idt, bt, cm, vd, desc in enums:
        rows.append([adt, idt, bt, "TRUE", cm, vd, desc])

    # ---- CS 内部用的基础类型（共享 boolean_CompuMethod，IDENTICAL 则无 CompuMethod） ----
    primitives = [
        ("ADT_Boolean", "IDT_Boolean", "boolean", "boolean_CompuMethod", "0:Invalid;1:Valid", "Boolean type (Invalid/Valid)"),
        ("ADT_OnOff", "IDT_OnOff", "boolean", "boolean_CompuMethod", "0:OFF;1:ON", "On/Off boolean"),
        ("ADT_ActiveFlag", "IDT_ActiveFlag", "boolean", "boolean_CompuMethod", "0:Inactive;1:Active", "Active/Inactive flag"),
        ("ADT_FaultFlag", "IDT_FaultFlag", "boolean", "boolean_CompuMethod", "0:Normal;1:Fault", "Fault status flag"),
        ("ADT_InhibitSts", "IDT_InhibitSts", "boolean", "", "", "Inhibit/disable status"),
    ]
    for adt, idt, bt, cm, vd, desc in primitives:
        rows.append([adt, idt, bt, "FALSE", cm, vd, desc])

    # ---- 外部 S/R 信号的通用占位类型（无 CompuMethod，只声明基础类型） ----
    rows.append(["ADT_Uint8", "IDT_Uint8", "uint8", "FALSE", "", "", "Generic uint8 placeholder for external S/R signals"])
    rows.append(["ADT_Uint16", "IDT_Uint16", "uint16", "FALSE", "", "", "Generic uint16 placeholder for external S/R signals"])

    return rows


# ============================================================================
# SR signal definitions
# ============================================================================

# Input signals: (interface_name, data_element_name, adt_type, description)
INPUT_SR_SIGNALS = [
    # ---- HWA hardware inputs (8) ----
    ("ifHWA_FWiperPark", "VbINP_HWA_FWiperPark_flg", "ADT_Boolean", "HWA front wiper Park position hardwire, filter 30ms"),
    ("ifHWA_RWiperPark", "VbINP_HWA_RWiperPark_flg", "ADT_Boolean", "HWA rear wiper Park position signal"),
    ("ifHWA_DCDCMode", "VbINP_HWA_DCDCModeSts_sig", "ADT_Uint8", "DC-DC converter mode"),
    ("ifHWA_FWiperMist", "VbINP_HWA_FWiperMistSts_flg", "ADT_Boolean", "Front wiper mist hardwire switch, T>=100ms debounce"),
    ("ifHWA_WasherLiquidLow", "VbINP_HWA_WasherLiquidLow_flg", "ADT_Boolean", "Washer liquid level sensor, 30s low=alarm, 6s recover"),
    ("ifHWA_IGNFeedBack", "VbINP_HWA_IGNFeedBackIN_flg", "ADT_Boolean", "IGN ignition feedback, Limphome enable condition"),
    ("ifHWA_RLSTimeOut", "VbINP_HWA_RLSTimeOut_flg", "ADT_Boolean", "RLS LIN timeout flag, triggers AUTO low-speed park stop"),
    ("ifHWA_Voltage", "VeINP_HWA_Voltage_100mV", "ADT_Uint16", "Vehicle voltage sample, out-of-range disables all wiper actions"),

    # ---- CAN front wiper/washer (10) ----
    ("ifCAN_ETRSFWiPerSw", "VeINP_CAN_ETRSFrontWiperSwitchStatus_sig", "ADT_Uint8", "ETRS front wiper lever position (OFF/Auto/Low/High)"),
    ("ifCAN_ETRSFWasherSw", "VeINP_CAN_ETRSFrontWasherSwitchStatus_sig", "ADT_Uint8", "ETRS front washer lever (OFF/MIST/FrontWasher)"),
    ("ifCAN_VoiceWashReq", "VeINP_CAN_ICMWiprWashVoiceReq_sig", "ADT_Uint8", "Voice wash request (event type), 3s timeout auto-off"),
    ("ifCAN_MaintenanceReq", "VbINP_CAN_ICMMaintenanceReq_flg", "ADT_Boolean", "Front wiper maintenance mode request, edge-triggered"),
    ("ifCAN_RainSensitivity", "VeINP_CAN_ICMRainSensitivity_sig", "ADT_Uint8", "ICM rain sensitivity setting, write EEPROM persistent"),
    ("ifCAN_RLSWiperSPD", "VeINP_CAN_RLSRQWiperSPD_sig", "ADT_Uint8", "RLS rain speed request V1 (reserved compat)"),
    ("ifCAN_RLSWiperSPDV2", "VeINP_CAN_RLSRQWiperSPDV2_sig", "ADT_Uint8", "RLS rain speed request V2 (primary), LIN->CAN route"),
    ("ifCAN_RLSFault", "VbINP_CAN_RLSFaultRain_flg", "ADT_Boolean", "RLS fault flag, AUTO stops after current cycle"),
    ("ifCAN_IDMWiperReq", "VeINP_CAN_IDMFwiperReq_sig", "ADT_Uint8", "IDM front wiper request, RLS priority over IDM on conflict"),
    ("ifCAN_WashModeSw", "VeINP_CAN_ICMWashModeSwSts_sig", "ADT_Uint8", "ICM wash mode switch (OFF/Fixed/Mobile)"),

    # ---- CAN rear / gear / heater / OTA (8) ----
    ("ifCAN_OTASts", "VeINP_CAN_ICMOTASts_sig", "ADT_Uint8", "OTA upgrade status, inhibits partial wiper actions"),
    ("ifCAN_RearWiperReq", "VeINP_CAN_ICMRearWiperReq_sig", "ADT_Uint8", "ICM rear wiper manual request, 0x2 triggers 3s intermittent"),
    ("ifCAN_RearWashSw", "VbINP_CAN_ICMRearWashSwSts_sig", "ADT_Uint8", "ICM rear wash switch, 3s auto-off, mutex with front wash"),
    ("ifCAN_RearMaintReq", "VeINP_CAN_ICMRearMaintenanceReq_flg", "ADT_Boolean", "Rear wiper maintenance request, edge-triggered"),
    ("ifCAN_HeatingReq", "VeINP_ICMFWindAndNozheatingReq_sig", "ADT_Uint8", "ICM windshield+nozzle heating request, >20min auto-off"),
    ("ifCAN_IDMHeater", "VeINP_CAN_IDMZCULHeater_sig", "ADT_Uint8", "IDM heater request, front nozzle only (no windshield)"),
    ("ifCAN_Gear", "VeINP_CAN_VCU1NActualGear_sig", "ADT_Uint8", "Actual gear, R rising edge triggers rear wiper; P/N limits IDM"),
    ("ifCAN_GearValid", "VbINP_CAN_VCU1FActualGear_flg", "ADT_Boolean", "Gear signal validity, must be co-valid with gear"),

    # ---- EPRM storage (3) ----
    ("ifEPRM_FMaintenance", "VbINP_EPRM_MaintenanceFromEE_flg", "ADT_Boolean", "EEPROM front maintenance flag restore on power-up"),
    ("ifEPRM_RMaintenance", "VbINP_EPRM_RearMaintenanceFromEE_flg", "ADT_Boolean", "EEPROM rear maintenance flag restore on power-up"),
    ("ifEPRM_RainSens", "VeINP_EPRM_ZCULRainSensitivityStsFromEE_sig", "ADT_Uint8", "EEPROM last rain sensitivity restore on power-up"),

    # ---- Vehicle status (7) ----
    ("ifVeh_PowerMode", "VeOUT_PDU_PowerMode_sig", "ADT_Uint8", "Vehicle power mode, non-OFF is wiper enable precond; OFF triggers park then sleep"),
    ("ifVeh_PowerModeValid", "VbOUT_PDU_PowerModeValid_flg", "ADT_Boolean", "Power mode validity, must co-valid with PowerMode"),
    ("ifVeh_PowerSource", "VeOUT_PDU_ZCULSystemPowerSource_sig", "ADT_Uint8", "System power source, affects Park homing logic"),
    ("ifVeh_CarMode", "VeOUT_CMS_ZCULCarMode_sig", "ADT_Uint8", "Vehicle working mode, wash/exhibition triggers WASHMODE=inhibit auto wiper"),
    ("ifVeh_WipingInhibit", "VbOUT_CMS_ZCULWipingInhibit_flg", "ADT_Boolean", "CMS auto wiper inhibit (one of WashMode conditions)"),
    ("ifVeh_AntiTheft", "VeOUT_ALM_ZCULAntiThelfSts_sig", "ADT_Uint8", "Anti-theft status, Armed/Alarm edge immediately cuts all wiper outputs"),
    ("ifVeh_Config", "VuINP_CFG_V23Type_sig", "ADT_Uint8", "Vehicle hardware config, distinguishes wiper motor variants"),
]

# Output signals per Atm: (interface_name, data_element_name, adt_type, atm_component, description)
OUTPUT_SR_PER_ATM = {
    "BOD_FWiper_Atm": [
        ("ifOUT_FWiperLow", "VbOUT_WW_FWiperLow_flg", "ADT_Boolean", "Front wiper low-speed motor output"),
        ("ifOUT_FWiperHigh", "VbOUT_WW_FWiperHigh_flg", "ADT_Boolean", "Front wiper high-speed motor output"),
    ],
    "BOD_FWasher_Atm": [
        ("ifOUT_FWsher", "VbOUT_WW_FWsher_flg", "ADT_Boolean", "Front washer motor output"),
    ],
    "BOD_RWiper_Atm": [
        ("ifOUT_RearFWiper", "VbOUT_WW_RearFWiper_flg", "ADT_Boolean", "Rear wiper motor output"),
    ],
    "BOD_RWasher_Atm": [
        ("ifOUT_RearFWsher", "VbOUT_WW_RearFWsher_flg", "ADT_Boolean", "Rear washer motor output"),
    ],
    "BOD_FWindHeater_Atm": [
        ("ifOUT_FWindHeater", "VbOUT_WW_FWindHeater_flg", "ADT_Boolean", "Front windshield heater motor output"),
    ],
    "BOD_FNozzleHeater_Atm": [
        ("ifOUT_FNozzleHeater", "VbOUT_WW_FNozzleHeater_flg", "ADT_Boolean", "Front nozzle heater motor output"),
    ],
    "BOD_WiperStatus_Atm": [
        ("ifOUT_FWiperSts", "VeOUT_WW_ZCULFWiperSts_sig", "ADT_Uint8", "Front wiper run status"),
        ("ifOUT_FWiperSWSts", "VeOUT_WW_ZCULFwiperSWSts_sig", "ADT_Uint8", "Front wiper switch status"),
        ("ifOUT_FWashingSts", "VbOUT_WW_ZCULFwiperwashingSts_flg", "ADT_Boolean", "Front washer status (OFF/ON)"),
        ("ifOUT_RWiperSts", "VeOUT_WW_ZCULRearWiperSts_flg", "ADT_Boolean", "Rear wiper status (OFF/ON)"),
        ("ifOUT_RWiperWashSts", "VeOUT_WW_ZCUL_RearWashWiperSts_flg", "ADT_Boolean", "Rear wash wiper status (OFF/ON)"),
        ("ifOUT_ParkPosition", "VbOUT_WW_ZCULParkPosition_flg", "ADT_Boolean", "Front wiper park position status"),
        ("ifOUT_AutoWipingInhibit", "VbOUT_WW_ZCULAutoWipingInhibit_flg", "ADT_Boolean", "Wash mode disable status"),
    ],
    "BOD_RainSensor_Atm": [
        ("ifOUT_RainSensSts", "VeOUT_WW_ZCULRainSensitivitySts_sig", "ADT_Uint8", "Rain sensor sensitivity status"),
        ("ifOUT_RainSensToEE", "VeOUT_WW_ZCULRainSensitivityStsToEE_sig", "ADT_Uint8", "Rain sensitivity write to E2ROM status"),
        ("ifOUT_RainSensorFail", "VbOUT_WW_ZCULRainSensorFailSts_flg", "ADT_Boolean", "Rain sensor fault status"),
    ],
    "BOD_Maintenance_Atm": [
        ("ifOUT_FMaintToEE", "VbOUT_WW_MaintenanceToEE_flg", "ADT_Boolean", "Front maintenance mode write to E2ROM status"),
        ("ifOUT_RMaintToEE", "VbOUT_WW_RearMaintenanceToEE_flg", "ADT_Boolean", "Rear maintenance mode write to E2ROM status"),
        ("ifOUT_RWMaintMode", "VeOUT_WW_ZCULRWMaintModeSts_flg", "ADT_Boolean", "Rear wiper maintenance mode (ON/OFF)"),
    ],
    "BOD_WashLiquid_Atm": [
        ("ifOUT_WashLiquidLow", "VbOUT_WW_ZCUL_WashingLiquidLow_flg", "ADT_Boolean", "Wash liquid low alarm status"),
    ],
    "BOD_HeaterStatus_Atm": [
        ("ifOUT_HeatingStsCAN", "VbOUT_CAN_ZCUL_FWindAndNozheatingSts_flg", "ADT_Boolean", "Front windshield + nozzle heating CAN status"),
    ],
}


# ============================================================================
# PortInterfaces = CS + SR
# ============================================================================
CS_INTERFACES = [
    ["rrFWiper", "CS", "", "", "FWiperCmd", "Front wiper execution service"],
    ["rrFWasher", "CS", "", "", "FWasherCmd", "Front washer drive service"],
    ["rrRWiper", "CS", "", "", "RWiperCmd", "Rear wiper execution service"],
    ["rrRWasher", "CS", "", "", "RWasherCmd", "Rear washer drive service"],
    ["rrFWindHeater", "CS", "", "", "FWindHeaterCmd", "Front windshield heater drive"],
    ["rrFNozzleHeater", "CS", "", "", "FNozzleHeaterCmd", "Front nozzle heater drive"],
    ["rrWiperStatus", "CS", "", "", "ReportWiperStatus", "Wiper/washer run status report"],
    ["rrRainSensor", "CS", "", "", "ReportRainSensor", "Rain sensor status & sensitivity report"],
    ["rrMaintenance", "CS", "", "", "ReportMaintenance", "Maintenance status report"],
    ["rrWashLiquid", "CS", "", "", "ReportWashLiquid", "Wash liquid alarm service"],
    ["rrHeaterStatus", "CS", "", "", "ReportHeaterStatus", "Heater status report"],
]


def build_port_interfaces():
    rows = list(CS_INTERFACES)

    # Input SR interfaces
    for if_name, de_name, adt, desc in INPUT_SR_SIGNALS:
        rows.append([if_name, "SR", de_name, adt, "", desc])

    # Output SR interfaces (dedup by interface_name)
    seen = set()
    for atm, signals in OUTPUT_SR_PER_ATM.items():
        for if_name, de_name, adt, desc in signals:
            if if_name not in seen:
                seen.add(if_name)
                rows.append([if_name, "SR", de_name, adt, "", desc])

    return rows


# ============================================================================
# Operations (unchanged CS logic)
# ============================================================================
def build_operations():
    rows = []

    # rrFWiper.FWiperCmd
    rows.append(["rrFWiper", "FWiperCmd", "FWiperCmd", "IN", "ADT_FWiperCmd", "Front wiper speed command (0:Stop,1:Low,2:High)"])
    rows.append(["rrFWiper", "FWiperCmd", "ReturnCode", "OUT", "ADT_ReturnCode", "Operation return code"])

    # rrFWasher.FWasherCmd
    rows.append(["rrFWasher", "FWasherCmd", "FWasherCmd", "IN", "ADT_Boolean", "Front washer ON/OFF"])
    rows.append(["rrFWasher", "FWasherCmd", "ReturnCode", "OUT", "ADT_ReturnCode", "Operation return code"])

    # rrRWiper.RWiperCmd
    rows.append(["rrRWiper", "RWiperCmd", "RWiperCmd", "IN", "ADT_Boolean", "Rear wiper ON/OFF"])
    rows.append(["rrRWiper", "RWiperCmd", "ReturnCode", "OUT", "ADT_ReturnCode", "Operation return code"])

    # rrRWasher.RWasherCmd
    rows.append(["rrRWasher", "RWasherCmd", "RWasherCmd", "IN", "ADT_Boolean", "Rear washer ON/OFF"])
    rows.append(["rrRWasher", "RWasherCmd", "ReturnCode", "OUT", "ADT_ReturnCode", "Operation return code"])

    # rrFWindHeater.FWindHeaterCmd
    rows.append(["rrFWindHeater", "FWindHeaterCmd", "FWindHeaterCmd", "IN", "ADT_Boolean", "Front windshield heater ON/OFF"])
    rows.append(["rrFWindHeater", "FWindHeaterCmd", "ReturnCode", "OUT", "ADT_ReturnCode", "Operation return code"])

    # rrFNozzleHeater.FNozzleHeaterCmd
    rows.append(["rrFNozzleHeater", "FNozzleHeaterCmd", "FNozzleHeaterCmd", "IN", "ADT_Boolean", "Front nozzle heater ON/OFF"])
    rows.append(["rrFNozzleHeater", "FNozzleHeaterCmd", "ReturnCode", "OUT", "ADT_ReturnCode", "Operation return code"])

    # rrWiperStatus.ReportWiperStatus (7 IN args + ReturnCode OUT)
    for name, adt, desc in [
        ("FWiperSts", "ADT_FWiperSts", "Front wiper run status"),
        ("FWiperSWSts", "ADT_FWiperSWSts", "Front wiper switch status"),
        ("FWashingSts", "ADT_OnOff", "Front washer status"),
        ("RWiperSts", "ADT_OnOff", "Rear wiper status"),
        ("RWiperWashSts", "ADT_OnOff", "Rear wash wiper status"),
        ("ParkPositionSts", "ADT_Boolean", "Front wiper park position"),
        ("AutoWipingInhibit", "ADT_InhibitSts", "Auto wiping inhibit"),
    ]:
        rows.append(["rrWiperStatus", "ReportWiperStatus", name, "IN", adt, desc])
    rows.append(["rrWiperStatus", "ReportWiperStatus", "ReturnCode", "OUT", "ADT_ReturnCode", "Operation return code"])

    # rrRainSensor.ReportRainSensor (3 IN + ReturnCode OUT)
    for name, adt, desc in [
        ("RainSensitivitySts", "ADT_RainSensitivity", "Rain sensitivity status"),
        ("RainSensitivityToEE", "ADT_RainSensitivity", "Rain sensitivity to E2ROM"),
        ("RainSensorFailSts", "ADT_FaultFlag", "Rain sensor fault (0:Normal,1:Fault)"),
    ]:
        rows.append(["rrRainSensor", "ReportRainSensor", name, "IN", adt, desc])
    rows.append(["rrRainSensor", "ReportRainSensor", "ReturnCode", "OUT", "ADT_ReturnCode", "Operation return code"])

    # rrMaintenance.ReportMaintenance (3 IN + ReturnCode OUT)
    for name, adt, desc in [
        ("FMaintenanceToEE", "ADT_ActiveFlag", "Front maintenance to E2ROM"),
        ("RMaintenanceToEE", "ADT_ActiveFlag", "Rear maintenance to E2ROM"),
        ("RWMaintModeSts", "ADT_OnOff", "Rear wiper maintenance mode"),
    ]:
        rows.append(["rrMaintenance", "ReportMaintenance", name, "IN", adt, desc])
    rows.append(["rrMaintenance", "ReportMaintenance", "ReturnCode", "OUT", "ADT_ReturnCode", "Operation return code"])

    # rrWashLiquid.ReportWashLiquid (1 IN + ReturnCode OUT)
    rows.append(["rrWashLiquid", "ReportWashLiquid", "WashingLiquidLow", "IN", "ADT_ActiveFlag", "Wash liquid low alarm"])
    rows.append(["rrWashLiquid", "ReportWashLiquid", "ReturnCode", "OUT", "ADT_ReturnCode", "Operation return code"])

    # rrHeaterStatus.ReportHeaterStatus (2 IN + ReturnCode OUT)
    for name, adt, desc in [
        ("FWindHeaterSts", "ADT_ActiveFlag", "Front windshield heater status"),
        ("FNozzleHeaterSts", "ADT_ActiveFlag", "Front nozzle heater status"),
    ]:
        rows.append(["rrHeaterStatus", "ReportHeaterStatus", name, "IN", adt, desc])
    rows.append(["rrHeaterStatus", "ReportHeaterStatus", "ReturnCode", "OUT", "ADT_ReturnCode", "Operation return code"])

    return rows


# ============================================================================
# Ports = CS + SR
# ============================================================================
def build_ports():
    rows = []

    # --- CS ports ---
    iface_to_atm = {
        "rrFWiper": "BOD_FWiper_Atm", "rrFWasher": "BOD_FWasher_Atm",
        "rrRWiper": "BOD_RWiper_Atm", "rrRWasher": "BOD_RWasher_Atm",
        "rrFWindHeater": "BOD_FWindHeater_Atm", "rrFNozzleHeater": "BOD_FNozzleHeater_Atm",
        "rrWiperStatus": "BOD_WiperStatus_Atm", "rrRainSensor": "BOD_RainSensor_Atm",
        "rrMaintenance": "BOD_Maintenance_Atm", "rrWashLiquid": "BOD_WashLiquid_Atm",
        "rrHeaterStatus": "BOD_HeaterStatus_Atm",
    }
    op_map = {
        "rrFWiper": "FWiperCmd", "rrFWasher": "FWasherCmd",
        "rrRWiper": "RWiperCmd", "rrRWasher": "RWasherCmd",
        "rrFWindHeater": "FWindHeaterCmd", "rrFNozzleHeater": "FNozzleHeaterCmd",
        "rrWiperStatus": "ReportWiperStatus", "rrRainSensor": "ReportRainSensor",
        "rrMaintenance": "ReportMaintenance", "rrWashLiquid": "ReportWashLiquid",
        "rrHeaterStatus": "ReportHeaterStatus",
    }

    # Enh R-Ports (CS client)
    for iface, atm in iface_to_atm.items():
        op = op_map[iface]
        rows.append(["WW_Enh", iface, "R", "CS", iface, "", op, "", "", f"Client calls {atm}"])

    # Atm P-Ports (CS server)
    for iface, atm in iface_to_atm.items():
        op = op_map[iface]
        rows.append([atm, iface, "P", "CS", iface, "", op, "", "", f"Server implements {iface}"])

    # --- SR input ports (Enh receives from external sources) ---
    for if_name, de_name, adt, desc in INPUT_SR_SIGNALS:
        port_name = "r" + if_name[4:]  # e.g. ifHWA_FWiperPark -> rHWA_FWiperPark
        rows.append(["WW_Enh", port_name, "R", "SR", if_name, de_name, "", "", "nonqueued", f"Receive: {desc}"])

    # --- SR output ports (Atm provides, Enh receives feedback) ---
    for atm, signals in OUTPUT_SR_PER_ATM.items():
        for if_name, de_name, adt, desc in signals:
            # Atm P-Port
            port_name = "p" + if_name[5:]  # e.g. ifOUT_FWiperLow -> pFWiperLow
            rows.append([atm, port_name, "P", "SR", if_name, de_name, "", "0", "nonqueued", f"Provide: {desc}"])
            # Enh R-Port (feedback)
            rport_name = "r" + if_name[5:]
            rows.append(["WW_Enh", rport_name, "R", "SR", if_name, de_name, "", "", "nonqueued", f"Receive feedback: {desc}"])

    return rows


# ============================================================================
# Runnables
# ============================================================================
def build_runnables():
    rows = [
        ["WW_Enh", "WW_Enh_Init", "WW_Enh_Init", "Initialization"],
        ["WW_Enh", "WW_Enh_MainTask", "WW_Enh_MainTask", "Main periodic control task (10ms)"],
    ]
    atm_names = [
        "BOD_FWiper_Atm", "BOD_FWasher_Atm", "BOD_RWiper_Atm", "BOD_RWasher_Atm",
        "BOD_FWindHeater_Atm", "BOD_FNozzleHeater_Atm", "BOD_WiperStatus_Atm",
        "BOD_RainSensor_Atm", "BOD_Maintenance_Atm", "BOD_WashLiquid_Atm", "BOD_HeaterStatus_Atm",
    ]
    op_names = [
        "FWiperCmd", "FWasherCmd", "RWiperCmd", "RWasherCmd",
        "FWindHeaterCmd", "FNozzleHeaterCmd", "ReportWiperStatus",
        "ReportRainSensor", "ReportMaintenance", "ReportWashLiquid", "ReportHeaterStatus",
    ]
    for atm, op in zip(atm_names, op_names):
        rows.append([atm, f"{atm}_Init", f"{atm}_Init", "Initialization"])
        rows.append([atm, op, op, f"Operation handler for {op}"])

    return rows


# ============================================================================
# RunnableEvents
# ============================================================================
def build_runnable_events():
    rows = []

    # ====== Enh: triggers ======
    rows.append(["WW_Enh", "WW_Enh_Init", "Init", "", "", "", "", "", "Power-on init"])
    rows.append(["WW_Enh", "WW_Enh_MainTask", "Periodic", "10", "", "", "", "", "10ms main control cycle"])

    # ====== Enh: access – Read (external input S/R R-Ports only) ======
    for if_name, de_name, adt, desc in INPUT_SR_SIGNALS:
        port_name = "r" + if_name[4:]
        rows.append(["WW_Enh", "WW_Enh_MainTask", "", "", port_name, "", "", "Read", f"Read: {desc}"])

    # ====== Enh: access – Call (all C/S R-Ports on Enh) ======
    cs_ifaces = [
        ("rrFWiper", "FWiperCmd"), ("rrFWasher", "FWasherCmd"),
        ("rrRWiper", "RWiperCmd"), ("rrRWasher", "RWasherCmd"),
        ("rrFWindHeater", "FWindHeaterCmd"), ("rrFNozzleHeater", "FNozzleHeaterCmd"),
        ("rrWiperStatus", "ReportWiperStatus"), ("rrRainSensor", "ReportRainSensor"),
        ("rrMaintenance", "ReportMaintenance"), ("rrWashLiquid", "ReportWashLiquid"),
        ("rrHeaterStatus", "ReportHeaterStatus"),
    ]
    for iface, op in cs_ifaces:
        rows.append(["WW_Enh", "WW_Enh_MainTask", "", "", iface, op, "", "Call", f"Call: {op} via {iface}"])

    # ====== Atm: triggers + access ======
    atm_op_cs_port = {
        "BOD_FWiper_Atm": ("rrFWiper", "FWiperCmd"),
        "BOD_FWasher_Atm": ("rrFWasher", "FWasherCmd"),
        "BOD_RWiper_Atm": ("rrRWiper", "RWiperCmd"),
        "BOD_RWasher_Atm": ("rrRWasher", "RWasherCmd"),
        "BOD_FWindHeater_Atm": ("rrFWindHeater", "FWindHeaterCmd"),
        "BOD_FNozzleHeater_Atm": ("rrFNozzleHeater", "FNozzleHeaterCmd"),
        "BOD_WiperStatus_Atm": ("rrWiperStatus", "ReportWiperStatus"),
        "BOD_RainSensor_Atm": ("rrRainSensor", "ReportRainSensor"),
        "BOD_Maintenance_Atm": ("rrMaintenance", "ReportMaintenance"),
        "BOD_WashLiquid_Atm": ("rrWashLiquid", "ReportWashLiquid"),
        "BOD_HeaterStatus_Atm": ("rrHeaterStatus", "ReportHeaterStatus"),
    }
    atm_op_sr_write_ports = {
        "BOD_FWiper_Atm": ["p_FWiperLow", "p_FWiperHigh"],
        "BOD_FWasher_Atm": ["p_FWsher"],
        "BOD_RWiper_Atm": ["p_RearFWiper"],
        "BOD_RWasher_Atm": ["p_RearFWsher"],
        "BOD_FWindHeater_Atm": ["p_FWindHeater"],
        "BOD_FNozzleHeater_Atm": ["p_FNozzleHeater"],
        "BOD_WiperStatus_Atm": [
            "p_FWiperSts", "p_FWiperSWSts", "p_FWashingSts", "p_RWiperSts",
            "p_RWiperWashSts", "p_ParkPosition", "p_AutoWipingInhibit",
        ],
        "BOD_RainSensor_Atm": ["p_RainSensSts", "p_RainSensToEE", "p_RainSensorFail"],
        "BOD_Maintenance_Atm": ["p_FMaintToEE", "p_RMaintToEE", "p_RWMaintMode"],
        "BOD_WashLiquid_Atm": ["p_WashLiquidLow"],
        "BOD_HeaterStatus_Atm": ["p_HeatingStsCAN"],
    }

    for atm, (cs_iface, op) in atm_op_cs_port.items():
        # Init trigger
        rows.append([atm, f"{atm}_Init", "Init", "", "", "", "", "", "Power-on init"])
        # OperationInvoked trigger – PortName = C/S P-Port (the trigger port)
        rows.append([atm, op, "OperationInvoked", "", cs_iface, op, "", "",
                     f"Triggered by client call on {cs_iface}"])
        # Write access – one row per S/R P-Port (data send)
        for sr_port in atm_op_sr_write_ports.get(atm, []):
            rows.append([atm, op, "", "", sr_port, op, "", "Write",
                         f"Write output via {sr_port}"])

    return rows


# ============================================================================
# CompositionConnectors = CS + SR
# ============================================================================
def build_composition_connectors():
    rows = []

    # --- CS assembly connectors ---
    cs_pairs = [
        ("BOD_FWiper_Atm", "rrFWiper"), ("BOD_FWasher_Atm", "rrFWasher"),
        ("BOD_RWiper_Atm", "rrRWiper"), ("BOD_RWasher_Atm", "rrRWasher"),
        ("BOD_FWindHeater_Atm", "rrFWindHeater"), ("BOD_FNozzleHeater_Atm", "rrFNozzleHeater"),
        ("BOD_WiperStatus_Atm", "rrWiperStatus"), ("BOD_RainSensor_Atm", "rrRainSensor"),
        ("BOD_Maintenance_Atm", "rrMaintenance"), ("BOD_WashLiquid_Atm", "rrWashLiquid"),
        ("BOD_HeaterStatus_Atm", "rrHeaterStatus"),
    ]
    for atm, iface in cs_pairs:
        rows.append(["WW_Comp", atm, iface, "WW_Enh", iface, "Assembly", f"CS: Enh -> {atm} via {iface}"])

    # --- SR assembly connectors (Atm P-Port -> Enh R-Port) ---
    for atm, signals in OUTPUT_SR_PER_ATM.items():
        for if_name, de_name, adt, desc in signals:
            p_port = "p" + if_name[5:]
            r_port = "r" + if_name[5:]
            rows.append(["WW_Comp", atm, p_port, "WW_Enh", r_port, "Assembly", f"SR feedback: {atm} -> Enh: {desc}"])

    return rows


# ============================================================================
# Validations & Write
# ============================================================================
def add_validations(workbook: Workbook) -> None:
    validations = {
        "Components": {"B": '"Application,Composition"', "D": '"TRUE,FALSE"'},
        "DataTypes": {"C": '"boolean,uint8,uint16,uint32,sint8,sint16,sint32,float32"', "D": '"TRUE,FALSE"'},
        "PortInterfaces": {"B": '"SR,CS"'},
        "Operations": {"D": '"IN,OUT,INOUT"'},
        "Ports": {"C": '"P,R"', "D": '"SR,CS"', "I": '"nonqueued,queued"'},
        "RunnableEvents": {
            "C": '"Init,Periodic,OperationInvoked,DataReceived"',
            "H": '"Read,Write,Call"',
        },
        "CompositionConnectors": {"F": '"Assembly,Delegation"'},
    }
    for sheet_name, column_rules in validations.items():
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for column, formula in column_rules.items():
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            sheet.add_data_validation(validation)
            validation.add(f"{column}2:{column}500")


def create_test_excel(path: Path) -> None:
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    title_font = Font(bold=True, size=18, name="Calibri")
    section_font = Font(bold=True, size=13, name="Calibri")
    meta_label_font = Font(bold=True, size=11, name="Calibri")
    text_font = Font(size=11, name="Calibri")

    data_funcs = {
        "Components": build_components,
        "DataTypes": build_data_types,
        "PortInterfaces": build_port_interfaces,
        "Operations": build_operations,
        "Ports": build_ports,
        "Runnables": build_runnables,
        "RunnableEvents": build_runnable_events,
        "CompositionConnectors": build_composition_connectors,
    }

    # ============================================================
    # Cover sheet — special rendering (no column headers)
    # ============================================================
    cover_sheet = workbook.create_sheet("Cover", 0)
    cover_rows = build_cover()
    row_idx = 1

    for item in cover_rows:
        row_type = item[0]
        if row_type == "title":
            cover_sheet.cell(row=row_idx, column=1, value=item[1]).font = title_font
            cover_sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            row_idx += 1
        elif row_type == "blank":
            row_idx += 1
        elif row_type == "meta":
            label_cell = cover_sheet.cell(row=row_idx, column=1, value=item[1])
            label_cell.font = meta_label_font
            cover_sheet.cell(row=row_idx, column=2, value=item[2]).font = text_font
            cover_sheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
            row_idx += 1
        elif row_type == "section":
            cover_sheet.cell(row=row_idx, column=1, value=item[1]).font = section_font
            cover_sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            row_idx += 1
        elif row_type == "text":
            for line_num, line in enumerate(item[1].split("\n")):
                cover_sheet.cell(row=row_idx + line_num, column=1, value=line).font = text_font
                cover_sheet.merge_cells(start_row=row_idx + line_num, start_column=1, end_row=row_idx + line_num, end_column=6)
            row_idx += item[1].count("\n") + 1
        elif row_type == "index_header":
            for ci, hdr in enumerate(["Sheet 名称", "说明"], start=1):
                c = cover_sheet.cell(row=row_idx, column=ci, value=hdr)
                c.font = Font(bold=True, size=11, name="Calibri")
                c.fill = PatternFill("solid", fgColor="D9EAF7")
            row_idx += 1
        elif row_type == "index_row":
            cover_sheet.cell(row=row_idx, column=1, value=item[1]).font = Font(bold=True, size=11, name="Calibri")
            cover_sheet.cell(row=row_idx, column=2, value=item[2]).font = text_font
            cover_sheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
            row_idx += 1

    cover_sheet.column_dimensions["A"].width = 22
    cover_sheet.column_dimensions["B"].width = 18
    for ci in range(3, 7):
        cover_sheet.column_dimensions[get_column_letter(ci)].width = 16

    # ============================================================
    # Business data sheets — standard header + data rendering
    # ============================================================
    for sheet_name, headers in SHEETS.items():
        if sheet_name == "Cover":
            continue
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        rows = data_funcs[sheet_name]()
        for row in rows:
            sheet.append(row)

        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            max_w = 14
            for cell in column:
                if cell.value:
                    max_w = max(max_w, min(len(str(cell.value)) + 2, 52))
            sheet.column_dimensions[column[0].column_letter].width = max_w

    add_validations(workbook)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    print(f"Created: {path}")
    stats = {k: len(v()) for k, v in data_funcs.items()}
    for k, v in stats.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    create_test_excel(OUTPUT)
